#!/usr/bin/env python3
# office_activity_simplifier_outlook_full_final_fixed.py
"""
Office Activity Simplifier - Full (fixed) version
Features:
 - Tasks (SQLite)
 - Contacts import (CSV/XLSX)
 - Responsible dropdown on task edit (from contacts)
 - Reminder HTML field (below progress Add entry)
 - "Send Reminder Now (Outlook)" button in edit window
 - Periodic reminder checker that fires popup and sends email if configured
"""

import calendar
import sys
import re
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import json
import os
import csv
import subprocess
import logging
import webbrowser
import urllib.parse
import urllib.request
import os
import subprocess
import requests
import msal
from datetime import datetime, date, timedelta

# ---- Persistent file logging (next to the app/exe) ----
import sys

try:
    # Directory where the script or frozen exe lives
    if getattr(sys, 'frozen', False):
        app_dir = os.path.dirname(sys.executable)
    else:
        app_dir = os.path.dirname(os.path.abspath(__file__))

    log_file = os.path.join(app_dir, "oats_debug.log")
    os.makedirs(app_dir, exist_ok=True)

    # Root logger file handler
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s %(levelname)s %(name)s %(message)s")
    fh.setFormatter(fmt)
    logging.getLogger().addHandler(fh)

    print(f"[DEBUG] Logging initialized. File: {log_file}")
except Exception as e:
    print(f"[WARN] Could not initialize file logging: {e}")

logger = logging.getLogger(__name__)
# enable console debug logging during development so we can see what's happening
import logging as _logging
if not _logging.getLogger().handlers:
    _logging.basicConfig(level=_logging.DEBUG, format="%(asctime)s %(levelname)s %(message)s")
logger.setLevel(_logging.DEBUG)
logger.addHandler(logging.NullHandler())

# Outlook availability (pywin32). On non-Windows machines this will be False.
try:
    import win32com.client  # type: ignore
    HAS_OUTLOOK = True
except Exception:
    HAS_OUTLOOK = False

# Safe import for win10toast
try:
    import importlib
    _win10toast = importlib.import_module("win10toast")
    def _create_toaster():
        try:
            return _win10toast.ToastNotifier()
        except Exception:
            return None
    toaster = _create_toaster()
    HAS_NOTIFY = toaster is not None
except Exception:
    toaster = None
    HAS_NOTIFY = False

# Optional HTML renderer in Kanban details
try:
    from tkhtmlview import HTMLLabel  # type: ignore
    HAS_HTML = True
except Exception:
    HAS_HTML = False

# optional calendar widget
try:
    from tkcalendar import DateEntry  # type: ignore
    HAS_DATEENTRY = True
except Exception:
    DateEntry = None
    HAS_DATEENTRY = False

DB_FILE = "office_tasks.db"
SETTINGS_FILE = "settings.json"

PRIORITIES = ["Low", "Medium", "High"]
STATUSES = ["Pending", "In-Progress", "Done"]


def _now_iso():
    return datetime.now().isoformat(timespec="seconds")


def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as f:
                return json.load(f)
        except Exception:
            logger.exception("Could not load settings.json")
    return {"outlook_refresh_minutes": 30, "show_description": False}


def save_settings(settings):
    with open(SETTINGS_FILE, "w") as f:
        json.dump(settings, f)


def _safe_show_toast(title, msg, duration=5):
    """
    Show a Windows toast if available. Catch and swallow all exceptions
    so toast library errors don't crash the app's callbacks.
    """
    global toaster, HAS_NOTIFY
    if not HAS_NOTIFY or toaster is None:
        return
    try:
        toaster.show_toast(title, (msg or "")[:200], duration=duration, threaded=True)
    except Exception:
        logger.exception("Toast error (ignored)")


# -------------------- Database --------------------
class TaskDB:
    def __init__(self, path=DB_FILE):
        self.conn = sqlite3.connect(path)
        # return rows as mapping
        self.conn.row_factory = sqlite3.Row
        # Improve durability / concurrency
        try:
            self.conn.execute("PRAGMA journal_mode=WAL;")
            self.conn.execute("PRAGMA foreign_keys=ON;")
        except Exception:
            pass
        self._init_db()

    def _init_db(self):
        cur = self.conn.cursor()
        # Primary tasks table
        cur.execute(
            """CREATE TABLE IF NOT EXISTS tasks(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                description TEXT,
                due_date TEXT,
                priority TEXT DEFAULT 'Medium',
                status TEXT DEFAULT 'Pending',
                created_at TEXT,
                updated_at TEXT,
                done_at TEXT,
                outlook_id TEXT,
                progress_log TEXT,
                attachments TEXT,
                reminder_minutes INTEGER,
                reminder_set_at TEXT,
                reminder_sent_at TEXT,
                deleted_at TEXT,
                recurrence TEXT,
                responsible_id INTEGER,
                reminder_email_body TEXT
            );"""
        )

        # contacts table for Name + Email
        cur.execute(
            """CREATE TABLE IF NOT EXISTS contacts(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                email TEXT,
                created_at TEXT
            );"""
        )

        # Ensure schema migrations (if db already exists but lacks these columns)
        existing_cols = set()
        try:
            cur.execute("PRAGMA table_info(tasks);")
            for row in cur.fetchall():
                # row is tuple-like: (cid, name, type, notnull, dflt_value, pk)
                try:
                    existing_cols.add(row["name"])
                except Exception:
                    # fallback if row is not a sqlite3.Row
                    existing_cols.add(row[1])
        except Exception:
            pass

        cols_to_ensure = [
            "outlook_id", "progress_log", "attachments", "reminder_minutes",
            "reminder_set_at", "reminder_sent_at", "deleted_at", "recurrence",
            "responsible_id", "reminder_email_body", "reminder_mail_entryid"
        ]
        for col in cols_to_ensure:
            try:
                if col not in existing_cols:
                    # text is fine for migration
                    cur.execute(f"ALTER TABLE tasks ADD COLUMN {col} TEXT;")
            except sqlite3.OperationalError:
                pass
            except Exception:
                pass

        self.conn.commit()

    # contact helpers
    def add_contact(self, name, email):
        now = _now_iso()
        with self.conn:
            self.conn.execute(
                "INSERT INTO contacts(name, email, created_at) VALUES(?,?,?)",
                (name, email, now)
            )

    def get_contacts(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id, name, email FROM contacts ORDER BY name")
        return cur.fetchall()

    def get_contact_label(self, contact_id):
        if not contact_id:
            return ""
        try:
            cur = self.conn.cursor()
            cur.execute("SELECT name, email FROM contacts WHERE id=?", (int(contact_id),))
            r = cur.fetchone()
            if not r:
                return ""
            name = r["name"] or ""
            email = r["email"] or ""
            return f"{name} <{email}>" if name else email
        except Exception:
            return ""

    def bulk_add_contacts_from_file(self, path):
        """
        Accepts a .csv or .xlsx file containing header columns 'name' and 'email'.
        Tries pandas if available for Excel; falls back to csv for .csv files or openpyxl fallback.
        """
        if not os.path.exists(path):
            return 0
        added = 0
        ext = os.path.splitext(path)[1].lower()
        rows = []
        try:
            if ext in (".xls", ".xlsx"):
                try:
                    import pandas as pd  # type: ignore
                    df = pd.read_excel(path)
                    for _, r in df.iterrows():
                        name = str(r.get("name") or r.get("Name") or "").strip()
                        email = str(r.get("email") or r.get("Email") or "").strip()
                        if email:
                            rows.append((name or email, email))
                except Exception:
                    # fallback: try openpyxl minimal read (no pandas)
                    try:
                        from openpyxl import load_workbook  # type: ignore
                        wb = load_workbook(path, read_only=True, data_only=True)
                        ws = wb.active
                        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.rows)]
                        name_idx = None
                        email_idx = None
                        for i, h in enumerate(headers):
                            if h in ("name", "full name", "fullname"):
                                name_idx = i
                            if h in ("email", "email id", "emailid", "email_address"):
                                email_idx = i
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            name = row[name_idx] if name_idx is not None else None
                            email = row[email_idx] if email_idx is not None else None
                            if email:
                                rows.append((str(name or "").strip() or str(email).strip(), str(email).strip()))
                    except Exception:
                        pass
            else:
                # csv fallback
                with open(path, newline="", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for r in reader:
                        name = (r.get("name") or r.get("Name") or "").strip()
                        email = (r.get("email") or r.get("Email") or "").strip()
                        if email:
                            rows.append((name or email, email))
        except Exception:
            pass

        # insert unique emails only
        cur = self.conn.cursor()
        for name, email in rows:
            try:
                cur.execute("SELECT 1 FROM contacts WHERE lower(email)=lower(?)", (email,))
                if cur.fetchone():
                    continue
                self.add_contact(name, email)
                added += 1
            except Exception:
                continue
        return added

    # tasks methods
    def add(self, title, description, due_date, priority, status="Pending", outlook_id=None,
            reminder_minutes=None, reminder_set_at=None, recurrence=None, responsible_id=None, reminder_email_body=None):
        now = _now_iso()
        done_at = now if status == "Done" else None
        with self.conn:
            self.conn.execute(
                """INSERT INTO tasks(title, description, due_date, priority, status,
                   created_at, updated_at, done_at, outlook_id, progress_log, reminder_minutes, reminder_set_at, reminder_sent_at, deleted_at, recurrence, responsible_id, reminder_email_body)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (title, description, due_date, priority, status, now, now, done_at, outlook_id, "", reminder_minutes, reminder_set_at, None, None, recurrence, responsible_id, reminder_email_body),
            )

    def update(self, task_id, title, description, due_date, priority, status, reminder_minutes=None, reminder_set_at=None, recurrence=None, responsible_id=None, reminder_email_body=None):
        now = _now_iso()
        done_at = now if status == "Done" else None
        with self.conn:
            # If caller passes reminder_* explicitly, update them; otherwise leave as-is
            if reminder_minutes is None and reminder_set_at is None and recurrence is None and responsible_id is None and reminder_email_body is None:
                self.conn.execute(
                    """UPDATE tasks SET title=?, description=?, due_date=?, priority=?, 
                       status=?, updated_at=?, done_at=? WHERE id=?""",
                    (title, description, due_date, priority, status, now, done_at, task_id),
                )
            else:
                self.conn.execute(
                    """UPDATE tasks SET title=?, description=?, due_date=?, priority=?, 
                       status=?, updated_at=?, done_at=?, reminder_minutes=?, reminder_set_at=?, recurrence=?, responsible_id=?, reminder_email_body=? WHERE id=?""",
                    (title, description, due_date, priority, status, now, done_at, reminder_minutes, reminder_set_at, recurrence, responsible_id, reminder_email_body, task_id),
                )

    def update_progress(self, task_id, progress_log):
        now = _now_iso()
        with self.conn:
            self.conn.execute(
                "UPDATE tasks SET progress_log=?, updated_at=? WHERE id=?",
                (progress_log, now, task_id),
            )

    def delete(self, task_id):
        with self.conn:
            self.conn.execute("DELETE FROM tasks WHERE id=?", (task_id,))

    def soft_delete(self, task_id):
        now = _now_iso()
        with self.conn:
            self.conn.execute("UPDATE tasks SET deleted_at=?, updated_at=? WHERE id=?", (now, now, task_id))

    def restore(self, task_id):
        with self.conn:
            self.conn.execute("UPDATE tasks SET deleted_at=NULL, updated_at=? WHERE id=?", (_now_iso(), task_id))

    def purge_deleted(self, older_than_iso=None):
        with self.conn:
            if older_than_iso:
                self.conn.execute("DELETE FROM tasks WHERE deleted_at IS NOT NULL AND deleted_at < ?", (older_than_iso,))
            else:
                self.conn.execute("DELETE FROM tasks WHERE deleted_at IS NOT NULL")

    def fetch(self):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE deleted_at IS NULL ORDER BY due_date IS NULL, due_date ASC, priority DESC")
        return cur.fetchall()

    def fetch_by_status(self, status):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE status=? AND deleted_at IS NULL ORDER BY priority DESC, due_date ASC", (status,))
        return cur.fetchall()

    def fetch_due_today(self):
        today = date.today().isoformat()
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE status!='Done' AND due_date=? AND deleted_at IS NULL ORDER BY priority DESC", (today,))
        return cur.fetchall()

    def fetch_overdue(self):
        today = date.today().isoformat()
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE status!='Done' AND due_date IS NOT NULL AND due_date < ? AND deleted_at IS NULL", (today,))
        return cur.fetchall()

    def fetch_deleted(self):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC")
        return cur.fetchall()

    def bulk_add(self, rows):
        now = _now_iso()
        with self.conn:
            for r in rows:
                done_at = now if r.get("status") == "Done" else None
                self.conn.execute(
                    """INSERT INTO tasks(title, description, due_date, priority, status,
                                        created_at, updated_at, done_at, outlook_id, progress_log, reminder_minutes, reminder_set_at, reminder_sent_at, deleted_at, recurrence)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        r.get("title"),
                        r.get("description"),
                        r.get("due_date"),
                        r.get("priority", "Medium"),
                        r.get("status", "Pending"),
                        now,
                        now,
                        done_at,
                        r.get("outlook_id"),
                        r.get("progress_log", ""),
                        None,
                        None,
                        None,
                        None,
                        r.get("recurrence")
                    ),
                )

    def mark_done(self, task_id):
        now = _now_iso()
        with self.conn:
            self.conn.execute(
                "UPDATE tasks SET status='Done', updated_at=?, done_at=? WHERE id=?",
                (now, now, task_id),
            )


# -------------------- App --------------------
class TaskApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Office Activity Simplifier")
        self.geometry("1400x850")
        self.db = TaskDB()
        self.settings = load_settings()

        # init style & theme
        self._init_styles()

        self.kanban_selected_id = None
        self.kanban_selected_status = None
        # mapping: status -> list of task_ids in same order as items in the listbox
        self.kanban_item_map = {status: [] for status in STATUSES}
        # mapping task_id -> card wrapper widget for Kanban visual selection
        self.kanban_card_widgets = {}
        # keep reference to currently highlighted widget (so we can un-highlight it)
        self._kanban_highlighted = None

        # attachments var
        self.attachments_var = tk.StringVar(value="")

        # Build UI
        self._build_ui()

        ### Global
        # Global debug binds — confirm any double-click reaches us
        try:
            # cheap test handler: will always log when any double-click happens anywhere
            def _dbg_any_double(e):
                logger.debug("GLOBAL DEBUG DOUBLE CLICK event: widget=%s x_root=%s y_root=%s type=%s",
                            getattr(e, "widget", None), getattr(e, "x_root", None), getattr(e, "y_root", None), getattr(e, "type", None))

            # bind the debug handler first (so you can see it in console)
            self.bind_all("<Double-Button-1>", _dbg_any_double, add="+")
            self.bind_all("<Double-1>", _dbg_any_double, add="+")

            # your real global handler (keep this)
            self.bind_all("<Double-Button-1>", self._global_kanban_double_click, add="+")
            self.bind_all("<Double-1>", self._global_kanban_double_click, add="+")
        except Exception:
            logger.exception("Failed to bind global double-click")

        


        # Key bindings
        try:
            self.tree.bind("<Delete>", lambda e: self._delete_task())
            #self.tree.bind("<BackSpace>", lambda e: self._delete_task())

            self.trash_tree.bind("<Delete>", lambda e: self._permanently_delete_selected_trash())
            #self.trash_tree.bind("<BackSpace>", lambda e: self._permanently_delete_selected_trash())

            for status, lb in self.kanban_lists.items():
                lb.bind("<Delete>", lambda e, _lb=lb: self._delete_selected_kanban())
                #lb.bind("<BackSpace>", lambda e, _lb=lb: self._delete_selected_kanban())

            self.bind_all("<Delete>", lambda e: self._on_delete_key())
            #self.bind_all("<BackSpace>", lambda e: self._on_delete_key())
        except Exception:
            pass

        self.bind_all("<Control-n>", lambda e: self._open_edit_window(None))
        self.bind_all("<Control-N>", lambda e: self._open_edit_window(None))

        # initial populate
        self.after(100, self._populate)
        self.after(100, self._populate_kanban)
        self.after(200, self._populate_trash)

        # reminders
        self._schedule_task_reminder_checker()
        self._refresh_reminder_display()

        if HAS_OUTLOOK:
            self._schedule_outlook_refresh(self.settings.get("outlook_refresh_minutes", 30))

        # check due-today toast hourly
        self._check_reminders()

        try:
            self.protocol("WM_DELETE_WINDOW", self._on_exit)
        except Exception:
            pass

    # ---------- Styles ----------
    def _init_styles(self):
        style = ttk.Style()
        try:
            default_font = ("Segoe UI", 10) if os.name == "nt" else ("Helvetica", 10)
            heading_font = ("Segoe UI", 11, "bold") if os.name == "nt" else ("Helvetica", 11, "bold")
            self.default_font = default_font
            self.heading_font = heading_font
            try:
                self.strike_font = (default_font[0], default_font[1], "overstrike")
            except Exception:
                self.strike_font = default_font

            style.configure(".", font=default_font)
            style.configure("Treeview.Heading", font=heading_font)
        except Exception:
            self.default_font = ("Helvetica", 10)
            self.heading_font = ("Helvetica", 11, "bold")
            self.strike_font = self.default_font

        self._themes = {
            "Light": {"bg": "#f7f7f7", "panel": "#ffffff", "kanban_bg": "#f0f0f0", "text": "#222222", "muted": "#666666"},
            "Dark": {"bg": "#2b2b2b", "panel": "#333333", "kanban_bg": "#3a3a3a", "text": "#ffffff", "muted": "#cccccc"}
        }
        self._current_theme = "Light"
        try:
            self.configure(bg=self._themes[self._current_theme]["bg"])
        except Exception:
            pass

    def _set_theme(self, theme_name):
        if theme_name not in self._themes:
            return
        self._current_theme = theme_name
        palette = self._themes[theme_name]
        try:
            self.configure(bg=palette["bg"])
        except Exception:
            pass
        self._populate()

    # -------------------- Reminder UI & Backend --------------------
    def _format_timedelta(self, td):
        total_seconds = int(td.total_seconds())
        if total_seconds <= 0:
            return "Now"
        hours, rem = divmod(total_seconds, 3600)
        minutes, seconds = divmod(rem, 60)
        if hours > 0:
            return f"{hours}h {minutes}m"
        if minutes >= 1:
            return f"{minutes}m {seconds}s"
        return f"{seconds}s"

    def _refresh_reminder_display(self):
        try:
            cur = self.db.conn.cursor()
            for iid in self.tree.get_children():
                vals = self.tree.item(iid, "values")
                if not vals:
                    continue
                try:
                    task_id = int(vals[0])
                except Exception:
                    continue

                cur.execute("SELECT reminder_minutes, reminder_set_at, reminder_sent_at FROM tasks WHERE id=?", (task_id,))
                row = cur.fetchone()
                display = "—"
                if row:
                    rm = row["reminder_minutes"]
                    set_at = row["reminder_set_at"]
                    sent_at = row["reminder_sent_at"]
                    if rm is None or rm == "" or set_at is None:
                        display = "—"
                    else:
                        try:
                            rm_int = int(rm)
                        except Exception:
                            display = "—"
                            self.tree.set(iid, "reminder", display)
                            continue
                        try:
                            set_dt = datetime.fromisoformat(set_at)
                        except Exception:
                            display = "—"
                            self.tree.set(iid, "reminder", display)
                            continue

                        target = set_dt + timedelta(minutes=rm_int)
                        if sent_at:
                            try:
                                sent_dt = datetime.fromisoformat(sent_at)
                                if sent_dt >= target:
                                    display = "Sent"
                                else:
                                    remaining = target - datetime.now()
                                    display = self._format_timedelta(remaining)
                            except Exception:
                                remaining = target - datetime.now()
                                display = self._format_timedelta(remaining)
                        else:
                            remaining = target - datetime.now()
                            display = self._format_timedelta(remaining)
                try:
                    self.tree.set(iid, "reminder", display)
                except Exception:
                    pass
        except Exception:
            pass
        finally:
            self.after(1000, self._refresh_reminder_display)

    def _schedule_task_reminder_checker(self):
        try:
            self._check_task_reminders()
        except Exception:
            logger.exception("Unhandled error in reminder scheduling")
        finally:
            try:
                self.after(5 * 1000, self._schedule_task_reminder_checker)
            except Exception:
                logger.exception("Failed to schedule next reminder check")

    def _check_task_reminders(self):
        """
        Check for tasks where reminder_set_at + reminder_minutes <= now and not yet reminder_sent_at.
        If a task has responsible_id and reminder_email_body, attempt to send.
        """
        try:
            cur = self.db.conn.cursor()
            cur.execute("""
                SELECT id, title, description, reminder_minutes, reminder_set_at, reminder_sent_at, responsible_id, reminder_email_body
                FROM tasks
                WHERE reminder_minutes IS NOT NULL AND reminder_minutes != '' AND reminder_set_at IS NOT NULL
                  AND status != 'Done'
            """)
            rows = cur.fetchall()
            now_dt = datetime.now()
            for r in rows:
                try:
                    rm_min = int(r["reminder_minutes"])
                except Exception:
                    continue
                try:
                    set_at = datetime.fromisoformat(r["reminder_set_at"])
                except Exception:
                    continue

                target = set_at + timedelta(minutes=rm_min)

                sent = None
                if r["reminder_sent_at"]:
                    try:
                        sent = datetime.fromisoformat(r["reminder_sent_at"])
                    except Exception:
                        sent = None

                if now_dt >= target and (sent is None or sent < target):
                    # show popup
                    self._show_reminder_popup(r["id"], r["title"], r["description"])

                    # attempt to send email if configured
                    try:
                        resp = r["responsible_id"]
                        body = r["reminder_email_body"]
                        title = r["title"] or "Task Reminder"
                        if resp and body:
                            # resolve email
                            try:
                                cro = self.db.conn.cursor()
                                cro.execute("SELECT email FROM contacts WHERE id=?", (int(resp),))
                                cro_r = cro.fetchone()
                                if cro_r and cro_r["email"]:
                                    to_email = cro_r["email"]
                                    self._send_reminder_email(int(r["id"]), to_email, title, body)
                            except Exception:
                                logger.exception("Could not resolve responsible email")
                    except Exception:
                        logger.exception("Error attempting to send reminder email")

                    # mark reminder_sent_at
                    now_iso = datetime.now().isoformat(timespec="seconds")
                    try:
                        self.db.conn.execute("UPDATE tasks SET reminder_sent_at=? WHERE id=?", (now_iso, r["id"]))
                        self.db.conn.commit()
                    except Exception:
                        logger.exception("Failed to update reminder_sent_at")
        except Exception:
            logger.exception("Reminder check error")

    def _show_reminder_popup(self, task_id, title, description):
        if HAS_NOTIFY:
            _safe_show_toast(f"Reminder: {title}", description or "Task due soon")

        win = tk.Toplevel(self)
        win.title("🔔 Task Reminder")
        win.geometry("640x320")
        try:
            if str(self.state()) == "iconic":
                self.deiconify()
        except Exception:
            pass
        try:
            win.attributes("-topmost", True)
            win.lift()
            win.update_idletasks()
            try:
                win.grab_set_global()
            except Exception:
                try:
                    win.grab_set()
                except Exception:
                    pass
            try:
                self.lift()
            except Exception:
                pass
        except Exception:
            pass

        header = ttk.Label(win, text=title, font=("", 14, "bold"))
        header.pack(padx=12, pady=(12, 6), anchor="w")

        txt = tk.Text(win, height=8, wrap="word", padx=8, pady=4)
        txt.insert("1.0", description or "(no description)")
        txt.config(state="disabled")
        txt.pack(fill=tk.BOTH, expand=False, padx=12, pady=(0, 8))

        btnf = ttk.Frame(win)
        btnf.pack(fill=tk.X, padx=12, pady=8)

        def open_task():
            try:
                win.grab_release()
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass
            try:
                self._open_edit_window(task_id)
            except Exception:
                logger.exception("Error opening task editor from reminder popup")

        def _snooze(minutes):
            try:
                minutes_int = int(minutes)
            except Exception:
                messagebox.showerror("Snooze Error", "Invalid snooze minutes value.")
                return

            new_set = datetime.now().isoformat(timespec="seconds")
            try:
                with self.db.conn:
                    self.db.conn.execute(
                        "UPDATE tasks SET reminder_minutes=?, reminder_set_at=?, reminder_sent_at=? WHERE id=?",
                        (minutes_int, new_set, None, task_id)
                    )
            except Exception:
                logger.exception("Snooze update error")
                try:
                    messagebox.showerror("Snooze Error", "Could not snooze reminder.", parent=win)
                except Exception:
                    pass
                return
            try:
                win.grab_release()
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass

        def dismiss():
            now_iso = datetime.now().isoformat(timespec="seconds")
            try:
                with self.db.conn:
                    self.db.conn.execute("UPDATE tasks SET reminder_sent_at=? WHERE id=?", (now_iso, task_id))
            except Exception:
                logger.exception("Dismiss update error")
                try:
                    messagebox.showerror("Dismiss Error", "Could not dismiss reminder.", parent=win)
                except Exception:
                    pass
            try:
                win.grab_release()
            except Exception:
                pass
            try:
                win.destroy()
            except Exception:
                pass

        ttk.Button(btnf, text="Open Task", command=open_task).pack(side=tk.LEFT, padx=6)
        ttk.Button(btnf, text="Snooze 5m", command=lambda: _snooze(5)).pack(side=tk.LEFT, padx=6)
        ttk.Button(btnf, text="Snooze 10m", command=lambda: _snooze(10)).pack(side=tk.LEFT, padx=6)
        ttk.Button(btnf, text="Snooze 30m", command=lambda: _snooze(30)).pack(side=tk.LEFT, padx=6)
        ttk.Button(btnf, text="Dismiss", command=dismiss).pack(side=tk.RIGHT, padx=6)

        try:
            win.focus_force()
            win.update()
            self.after(1500, lambda: win.attributes("-topmost", True))
        except Exception:
            pass



    ###
    def _create_filter_bar(self, parent):
        """
        Creates the shared filter bar in the given parent frame.
        Uses the same self.filter_* variables used by Task List so filters are global.
        """
        # Ensure filter variables exist (they are created for Task List normally; create if missing)
        if not hasattr(self, "filter_text_var"):
            self.filter_text_var = tk.StringVar(value="")
        if not hasattr(self, "filter_priority_var"):
            self.filter_priority_var = tk.StringVar(value="All")
        if not hasattr(self, "filter_status_var"):
            self.filter_status_var = tk.StringVar(value="All")
        if not hasattr(self, "filter_show_completed_var"):
            self.filter_show_completed_var = tk.StringVar(value="Yes")
        if not hasattr(self, "filter_due_var"):
            self.filter_due_var = tk.StringVar(value="")

        filter_frame = ttk.Frame(parent, padding=(6, 4))
        filter_frame.pack(fill=tk.X, padx=6, pady=(6, 4))

        ttk.Label(filter_frame, text="Search:").pack(side=tk.LEFT, padx=(0, 4))
        search_entry = ttk.Entry(filter_frame, textvariable=self.filter_text_var, width=30)
        search_entry.pack(side=tk.LEFT)
        search_entry.bind("<KeyRelease>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Priority:").pack(side=tk.LEFT, padx=(12, 4))
        pri_vals = ["All"] + PRIORITIES
        pri_cb = ttk.Combobox(filter_frame, textvariable=self.filter_priority_var, values=pri_vals, width=10, state="readonly")
        pri_cb.pack(side=tk.LEFT)
        pri_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Status:").pack(side=tk.LEFT, padx=(12, 4))
        stat_vals = ["All"] + STATUSES
        stat_cb = ttk.Combobox(filter_frame, textvariable=self.filter_status_var, values=stat_vals, width=12, state="readonly")
        stat_cb.pack(side=tk.LEFT)
        stat_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Show Completed:").pack(side=tk.LEFT, padx=(12, 4))
        show_vals = ["Yes", "No"]
        show_cb = ttk.Combobox(filter_frame, textvariable=self.filter_show_completed_var, values=show_vals, width=6, state="readonly")
        show_cb.pack(side=tk.LEFT)
        show_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Due on (YYYY-MM-DD):").pack(side=tk.LEFT, padx=(12, 4))
        due_entry = ttk.Entry(filter_frame, textvariable=self.filter_due_var, width=12)
        due_entry.pack(side=tk.LEFT)

        ttk.Button(filter_frame, text="Apply", command=self._apply_filters).pack(side=tk.LEFT, padx=(12, 4))
        ttk.Button(filter_frame, text="Clear", command=self._clear_filters).pack(side=tk.LEFT)

        return filter_frame

    # -------------------- UI / CRUD / Kanban --------------------
    def _build_ui(self):
        toolbar = ttk.Frame(self, padding=8)
        toolbar.pack(fill=tk.X)
        ttk.Button(toolbar, text="Import Outlook Tasks", command=self._import_outlook_flags).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="Refresh Outlook", command=self._refresh_outlook_flags).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="Import CSV", command=self._import_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="Export CSV", command=self._export_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="Show Overdue", command=self._show_overdue_popup).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="Show Today", command=self._show_today_popup).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="Import Contacts", command=self._import_contacts).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="Settings", command=self._open_settings).pack(side=tk.RIGHT, padx=5)

        ttk.Label(toolbar, text="Theme:").pack(side=tk.RIGHT, padx=(6, 4))
        self.theme_var = tk.StringVar(value=self._current_theme)
        theme_cb = ttk.Combobox(toolbar, textvariable=self.theme_var, values=list(self._themes.keys()), width=10, state="readonly")
        theme_cb.pack(side=tk.RIGHT, padx=(0, 8))
        theme_cb.bind("<<ComboboxSelected>>", lambda e: self._set_theme(self.theme_var.get()))

        # Main notebook
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        # Task List tab
        list_tab = ttk.Frame(self.notebook)
        self.notebook.add(list_tab, text="Task List")

        # include 'reminder' column always and 'responsible'
                # include 'reminder' column always and 'responsible'
        # Make Task ID visible in the task list
        if self.settings.get("show_description", False):
            cols = ["id", "title", "desc", "due", "priority", "status", "responsible", "reminder"]
            # include id in display cols so it's visible
            display_cols = ["id", "title", "desc", "due", "priority", "status", "responsible", "reminder"]
        else:
            cols = ["id", "title", "due", "priority", "status", "responsible", "reminder"]
            display_cols = ["id", "title", "due", "priority", "status", "responsible", "reminder"]

        # Filter bar
        #filter_frame = ttk.Frame(list_tab, padding=(6, 4))

        # Reusable filter bar (Task List)
        self._create_filter_bar(list_tab)
        """/* filter_frame.pack(fill=tk.X, padx=6, pady=(6, 4))

        ttk.Label(filter_frame, text="Search:").pack(side=tk.LEFT, padx=(0, 4))
        self.filter_text_var = tk.StringVar(value="")
        search_entry = ttk.Entry(filter_frame, textvariable=self.filter_text_var, width=30)
        search_entry.pack(side=tk.LEFT)
        search_entry.bind("<KeyRelease>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Priority:").pack(side=tk.LEFT, padx=(12, 4))
        self.filter_priority_var = tk.StringVar(value="All")
        pri_vals = ["All"] + PRIORITIES
        pri_cb = ttk.Combobox(filter_frame, textvariable=self.filter_priority_var, values=pri_vals, width=10, state="readonly")
        pri_cb.pack(side=tk.LEFT)
        pri_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Status:").pack(side=tk.LEFT, padx=(12, 4))
        self.filter_status_var = tk.StringVar(value="All")
        stat_vals = ["All"] + STATUSES
        stat_cb = ttk.Combobox(filter_frame, textvariable=self.filter_status_var, values=stat_vals, width=12, state="readonly")
        stat_cb.pack(side=tk.LEFT)
        stat_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Show Completed:").pack(side=tk.LEFT, padx=(12, 4))
        self.filter_show_completed_var = tk.StringVar(value="Yes")
        show_vals = ["Yes", "No"]
        show_cb = ttk.Combobox(filter_frame, textvariable=self.filter_show_completed_var, values=show_vals, width=6, state="readonly")
        show_cb.pack(side=tk.LEFT)
        show_cb.bind("<<ComboboxSelected>>", lambda e: self._apply_filters())

        ttk.Label(filter_frame, text="Due on (YYYY-MM-DD):").pack(side=tk.LEFT, padx=(12, 4))
        self.filter_due_var = tk.StringVar(value="")
        due_entry = ttk.Entry(filter_frame, textvariable=self.filter_due_var, width=12)
        due_entry.pack(side=tk.LEFT)

        ttk.Button(filter_frame, text="Apply", command=self._apply_filters).pack(side=tk.LEFT, padx=(12, 4))
        ttk.Button(filter_frame, text="Clear", command=self._clear_filters).pack(side=tk.LEFT)
        """
        # Treeview
        self.tree = ttk.Treeview(list_tab, columns=cols, show="headings", displaycolumns=display_cols)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        for col in cols:
            header_text = col.title() if col != "id" else "ID"
            self.tree.heading(col, text=header_text, command=lambda _col=col: self._treeview_sort_column(_col, False))

        # Column widths
        if self.settings.get("show_description", False):
            self.tree.column("id", width=60, anchor="center")
            self.tree.column("title", width=300, anchor="w")
            self.tree.column("desc", width=300, anchor="w")
            self.tree.column("due", width=100, anchor="center")
            self.tree.column("priority", width=90, anchor="center")
            self.tree.column("status", width=90, anchor="center")
            self.tree.column("responsible", width=180, anchor="w")
            self.tree.column("reminder", width=120, anchor="center")
        else:
            self.tree.column("id", width=60, anchor="center")
            self.tree.column("title", width=420, anchor="w")
            self.tree.column("due", width=120, anchor="center")
            self.tree.column("priority", width=100, anchor="center")
            self.tree.column("status", width=100, anchor="center")
            self.tree.column("responsible", width=180, anchor="w")
            self.tree.column("reminder", width=120, anchor="center")

        try:
            self.tree.tag_configure("priority_high", background="#FFD6D6")
            self.tree.tag_configure("priority_medium", background="#FFF5CC")
            self.tree.tag_configure("priority_low", background="#E6FFEA")
            self.tree.tag_configure("oddrow", background="#FFFFFF")
            self.tree.tag_configure("evenrow", background="#F6F6F6")
        except Exception:
            pass

        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Double-1>", self._on_task_double_click)

        # Buttons
        btns = ttk.Frame(list_tab)
        btns.pack(fill=tk.X, padx=8, pady=5)
        ttk.Button(btns, text="Add", command=lambda: self._open_edit_window(None)).pack(side=tk.LEFT)
        ttk.Button(btns, text="Edit", command=lambda: self._open_edit_window(self._selected_tree_task_id() or None)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="Mark Done", command=self._mark_done).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="Delete", command=self._delete_task).pack(side=tk.LEFT, padx=5)

        # Kanban tab
        self.kanban_tab = ttk.Frame(self.notebook, padding=10)
                # Add the shared filter bar to Kanban
        self._create_filter_bar(self.kanban_tab)
        self.notebook.add(self.kanban_tab, text="Kanban Board")

        frame = ttk.Frame(self.kanban_tab)
        frame.pack(fill=tk.BOTH, expand=True)

                # Kanban columns - improved card UI
        self.kanban_columns = {}  # status -> dict(canvas, inner_frame, scrollbar)
        for idx, status in enumerate(STATUSES):
            col = ttk.Frame(frame, padding=6, borderwidth=1, relief="groove")
            col.grid(row=0, column=idx, sticky="nsew", padx=6)
            frame.columnconfigure(idx, weight=2)

            ttk.Label(col, text=status, font=("", 12, "bold")).pack(anchor="w")

            # Canvas to hold cards + inner frame
            canvas_col = tk.Canvas(col, height=700, highlightthickness=0)
            vscroll = ttk.Scrollbar(col, orient=tk.VERTICAL, command=canvas_col.yview)
            canvas_col.configure(yscrollcommand=vscroll.set)
            vscroll.pack(side=tk.RIGHT, fill=tk.Y)
            canvas_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            inner = ttk.Frame(canvas_col)
            canvas_col.create_window((0, 0), window=inner, anchor="nw")

            def _configure_inner(e, c=canvas_col):
                c.configure(scrollregion=c.bbox("all"))
            inner.bind("<Configure>", _configure_inner)

            # store references
            self.kanban_columns[status] = {"canvas": canvas_col, "frame": inner, "scroll": vscroll, "status_name": status}

            # clicking on the canvas/frame will select card via bind (cards will be clickable)

        # Kanban details panel
        desc_frame = ttk.Frame(frame, padding=6, borderwidth=1, relief="groove")
        desc_frame.grid(row=0, column=len(STATUSES), sticky="nsew", padx=6)
        frame.columnconfigure(len(STATUSES), weight=1)

        if HAS_HTML:
            self.kanban_html = HTMLLabel(desc_frame, html="", width=50, height=15)
            self.kanban_text = tk.Text(desc_frame, wrap="word", height=15, width=50)
        else:
            self.kanban_html = tk.Text(desc_frame, wrap="word", height=15, width=50)
            self.kanban_text = self.kanban_html

        ttk.Label(desc_frame, text="Task Description / Email").pack(anchor="w")
        self.kanban_text.pack(fill=tk.BOTH, expand=True)

        self.btn_save_desc = ttk.Button(desc_frame, text="Save Description", command=self._save_kanban_desc)
        self.btn_save_desc.pack(pady=5)

        ttk.Label(desc_frame, text="Progress Log").pack(anchor="w")
        self.kanban_progress = tk.Text(desc_frame, height=8, wrap="word", width=50)
        self.kanban_progress.pack(fill=tk.BOTH, expand=True)
        ttk.Button(desc_frame, text="Update Progress", command=self._update_progress).pack(pady=5)

        ttk.Label(desc_frame, text="Attachments").pack(anchor="w", pady=(10, 0))
        self.kanban_attachments_var = tk.StringVar()
        self.kanban_attachments_label = ttk.Label(desc_frame, textvariable=self.kanban_attachments_var, wraplength=350)
        self.kanban_attachments_label.pack(anchor="w", fill=tk.X, pady=2)
        ttk.Button(desc_frame, text="Open Attachments", command=self._open_selected_kanban_attachments).pack(anchor="w", pady=2)

        # Trash tab
        trash_tab = ttk.Frame(self.notebook)
                # Add the shared filter bar to Trash
        self._create_filter_bar(trash_tab)
        self.notebook.add(trash_tab, text="Trash")

        trash_toolbar = ttk.Frame(trash_tab, padding=6)
        trash_toolbar.pack(fill=tk.X)
        ttk.Button(trash_toolbar, text="Restore", command=self._restore_selected_trash).pack(side=tk.LEFT, padx=4)
        ttk.Button(trash_toolbar, text="Delete Permanently", command=self._permanently_delete_selected_trash).pack(side=tk.LEFT, padx=4)
        ttk.Button(trash_toolbar, text="Empty Trash", command=self._empty_trash_confirm).pack(side=tk.LEFT, padx=4)
        ttk.Button(trash_toolbar, text="Refresh", command=self._populate_trash).pack(side=tk.RIGHT, padx=4)

        self.trash_tree = ttk.Treeview(trash_tab, columns=["id", "title", "deleted_at", "due", "priority", "status"], show="headings")
        self.trash_tree.heading("id", text="ID")
        self.trash_tree.heading("title", text="Title")
        self.trash_tree.heading("deleted_at", text="Deleted At")
        self.trash_tree.heading("due", text="Due")
        self.trash_tree.heading("priority", text="Priority")
        self.trash_tree.heading("status", text="Status")
        self.trash_tree.column("id", width=60, anchor="center")
        self.trash_tree.column("title", width=400, anchor="w")
        self.trash_tree.column("deleted_at", width=160, anchor="center")
        self.trash_tree.column("due", width=120, anchor="center")
        self.trash_tree.column("priority", width=90, anchor="center")
        self.trash_tree.column("status", width=90, anchor="center")
        self.trash_tree.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        self.trash_tree.bind("<Double-1>", lambda e: self._restore_selected_trash())

        action_frame = ttk.Frame(self.kanban_tab, padding=5)
        action_frame.pack(fill=tk.X)
        self.btn_edit = ttk.Button(action_frame, text="Edit", command=self._edit_selected_kanban, state="disabled"); self.btn_edit.pack(side=tk.LEFT, padx=5)
        self.btn_delete = ttk.Button(action_frame, text="Delete", command=self._delete_selected_kanban, state="disabled"); self.btn_delete.pack(side=tk.LEFT, padx=5)
        self.btn_done = ttk.Button(action_frame, text="Mark Done", command=self._mark_done_selected_kanban, state="disabled"); self.btn_done.pack(side=tk.LEFT, padx=5)
        self.btn_prev = ttk.Button(action_frame, text="← Move Previous", command=self._move_prev_selected, state="disabled"); self.btn_prev.pack(side=tk.LEFT, padx=5)
        self.btn_next = ttk.Button(action_frame, text="Move Next →", command=self._move_next_selected, state="disabled"); self.btn_next.pack(side=tk.LEFT, padx=5)

    # -------------------- Edit Window --------------------
    def _open_edit_window(self, task_id=None):
        """
        Open popup for adding/editing a task.
        Reminder Email field is placed below the progress 'Add entry' area,
        and a 'Send Reminder Now (Outlook)' button is present.
        """
        try:
            from tkcalendar import DateEntry  # type: ignore
            has_dateentry = True
        except Exception:
            DateEntry = None
            has_dateentry = False

        win = tk.Toplevel(self)
        win.transient(self)
        win.title("Edit Task" if task_id else "Add Task")
        try:
            win.grab_set()
        except Exception:
            pass

        # size
        try:
            screen_w = self.winfo_screenwidth()
            screen_h = self.winfo_screenheight()
            width = min(1000, int(screen_w * 0.85))
            height = min(900, int(screen_h * 0.85))
            min_width, min_height = 700, 520
            width = max(min_width, width)
            height = max(min_height, height)
            win.geometry(f"{width}x{height}")
            win.minsize(min_width, min_height)
            win.resizable(True, True)
        except Exception:
            try:
                win.geometry("900x700")
                win.minsize(700, 520)
            except Exception:
                pass

        # container + canvas for scroll
        container = ttk.Frame(win)
        container.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(container, highlightthickness=0)
        vscroll = ttk.Scrollbar(container, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        content_frame = ttk.Frame(canvas, padding=10)
        canvas.create_window((0, 0), window=content_frame, anchor="nw")

        def _on_frame_configure(event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
        content_frame.bind("<Configure>", _on_frame_configure)

        def _on_mousewheel(event):
            if sys.platform.startswith("win") or sys.platform.startswith("linux"):
                delta = -1 * int(event.delta / 120)
            else:
                delta = -1 * int(event.delta)
            canvas.yview_scroll(delta, "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        # columns
        for c in range(6):
            content_frame.columnconfigure(c, weight=0)
        content_frame.columnconfigure(1, weight=1)

        # variables
        title_var = tk.StringVar()
        due_var = tk.StringVar()
        priority_var = tk.StringVar(value="Medium")
        status_var = tk.StringVar(value="Pending")
        reminder_var = tk.StringVar(value="")
        rec_type_var = tk.StringVar(value="None")
        rec_n_var = tk.StringVar(value="1")

        staged_attachments = []
        existing_attachments = []
        staged_progress_entries = ""

        row = 0
        # Title
        ttk.Label(content_frame, text="Title *").grid(row=row, column=0, sticky="w")
        ttk.Entry(content_frame, textvariable=title_var, width=60).grid(row=row, column=1, columnspan=4, sticky="we", padx=6, pady=4)
        row += 1

        # Due date
        ttk.Label(content_frame, text="Due Date (YYYY-MM-DD)").grid(row=row, column=0, sticky="w")
        if has_dateentry:
            DateEntry(content_frame, date_pattern="yyyy-mm-dd", textvariable=due_var, width=20).grid(row=row, column=1, sticky="w", padx=6, pady=4)
        else:
            ttk.Entry(content_frame, textvariable=due_var, width=20).grid(row=row, column=1, sticky="w", padx=6, pady=4)

        # Priority & Status
        ttk.Label(content_frame, text="Priority").grid(row=row, column=2, sticky="w", padx=(12, 0))
        ttk.Combobox(content_frame, textvariable=priority_var, values=PRIORITIES, state="readonly", width=12).grid(row=row, column=3, sticky="w", padx=6, pady=4)
        ttk.Label(content_frame, text="Status").grid(row=row, column=4, sticky="w", padx=(12, 0))
        ttk.Combobox(content_frame, textvariable=status_var, values=STATUSES, state="readonly", width=14).grid(row=row, column=5, sticky="w", padx=6, pady=4)
        row += 1

        # Reminder + Recurrence row
        ttk.Label(content_frame, text="Reminder (minutes before due)").grid(row=row, column=0, sticky="w")
        reminder_choices = ["", "5", "10", "30", "60", "120", "1440"]
        reminder_cb = ttk.Combobox(content_frame, textvariable=reminder_var, values=reminder_choices, width=18)
        reminder_cb.grid(row=row, column=1, sticky="w", padx=6, pady=4)
        reminder_cb.set("")

        ttk.Label(content_frame, text="Recurrence").grid(row=row, column=2, sticky="w", padx=(12, 0))
        rec_types = ["None", "Every N days", "Every N weeks", "Every N months"]
        rec_cb = ttk.Combobox(content_frame, textvariable=rec_type_var, values=rec_types, state="readonly", width=18)
        rec_cb.grid(row=row, column=3, sticky="w", padx=6, pady=4)
        ttk.Label(content_frame, text="N").grid(row=row, column=4, sticky="w")
        ttk.Entry(content_frame, textvariable=rec_n_var, width=6).grid(row=row, column=5, sticky="w", padx=(4, 0))
        row += 1

        # Responsible (dropdown) - choices will be set from DB
        # Responsible (dropdown populated from contacts)
        ttk.Label(content_frame, text="Responsible (contact)").grid(row=row, column=0, sticky="w", pady=(12,0))
        responsible_var = tk.StringVar(value="")
        responsible_cb = ttk.Combobox(content_frame, textvariable=responsible_var, values=[], width=40, state="readonly")
        responsible_cb.grid(row=row, column=1, columnspan=2, sticky="w", padx=6, pady=(8,0))
        row += 1

        # Reminder email body (HTML) and toolbar
        email_toolbar = ttk.Frame(content_frame)
        email_toolbar.grid(row=row, column=1, columnspan=4, sticky="w", padx=6, pady=(6, 0))
        # Helper: load contacts to combobox (callable so we can refresh after import)
        def _load_contacts_to_combobox():
            try:
                contacts = self.db.get_contacts()
                choices = []
                lookup_map = {}
                for c in contacts:
                    cid = c["id"]
                    name = c["name"] or ""
                    email = c["email"] or ""
                    label = f"{name} <{email}>" if name else email
                    choices.append(label)
                    lookup_map[label] = cid
                responsible_cb['values'] = choices
                responsible_cb.lookup_map = lookup_map
            except Exception:
                responsible_cb['values'] = []
                responsible_cb.lookup_map = {}

        # Insert HTML building blocks into the email body text widget
        def _insert_html_at_cursor(html_snippet):
            try:
                pos = email_body_text.index(tk.INSERT)
                email_body_text.insert(pos, html_snippet)
                email_body_text.focus_set()
            except Exception:
                try:
                    email_body_text.insert(tk.END, html_snippet)
                except Exception:
                    pass

        def _insert_table_dialog():
            dlg = tk.Toplevel(win)
            dlg.title("Insert Table")
            ttk.Label(dlg, text="Rows:").grid(row=0, column=0, padx=6, pady=6)
            rows_var = tk.IntVar(value=2)
            ttk.Entry(dlg, textvariable=rows_var, width=6).grid(row=0, column=1, padx=6, pady=6)
            ttk.Label(dlg, text="Cols:").grid(row=1, column=0, padx=6, pady=6)
            cols_var = tk.IntVar(value=2)
            ttk.Entry(dlg, textvariable=cols_var, width=6).grid(row=1, column=1, padx=6, pady=6)

            def _do_insert_table():
                r = max(1, int(rows_var.get()))
                c = max(1, int(cols_var.get()))
                html = "<table border='1' cellpadding='4' cellspacing='0'>\n"
                for _ in range(r):
                    html += "  <tr>\n"
                    for _ in range(c):
                        html += "    <td>&nbsp;</td>\n"
                    html += "  </tr>\n"
                html += "</table>\n"
                _insert_html_at_cursor(html)
                try:
                    dlg.destroy()
                except Exception:
                    pass

            ttk.Button(dlg, text="Insert", command=_do_insert_table).grid(row=2, column=0, columnspan=2, pady=8)

        def _insert_image_dialog():
            f = filedialog.askopenfilename(parent=win, filetypes=[("Image files", "*.png;*.jpg;*.jpeg;*.gif;*.bmp"), ("All files","*.*")])
            if not f:
                return
            file_url = f"file://{os.path.abspath(f)}"
            html = f'<img src="{file_url}" alt="Image" style="max-width:600px; height:auto;" />\n'
            _insert_html_at_cursor(html)

        def _insert_snippet():
            dlg = tk.Toplevel(win)
            dlg.title("Insert HTML Snippet")
            txt = tk.Text(dlg, height=8, width=60, wrap="none")
            txt.pack(padx=6, pady=6)
            def _ok():
                snippet = txt.get("1.0", tk.END)
                _insert_html_at_cursor(snippet)
                try:
                    dlg.destroy()
                except Exception:
                    pass
            ttk.Button(dlg, text="Insert", command=_ok).pack(pady=(0,8))

        ttk.Button(email_toolbar, text="Table", command=_insert_table_dialog).pack(side=tk.LEFT, padx=(0,4))
        ttk.Button(email_toolbar, text="Image", command=_insert_image_dialog).pack(side=tk.LEFT, padx=(0,4))
        ttk.Button(email_toolbar, text="Snippet", command=_insert_snippet).pack(side=tk.LEFT, padx=(0,4))
        _load_contacts_to_combobox()
        row += 1


        # If there is a responsible_id stored, map it to the combobox label


        # Description (below)
        ttk.Label(content_frame, text="Description").grid(row=row, column=0, sticky="nw", pady=(6, 0))
        desc_text = tk.Text(content_frame, height=8, width=80, wrap="word")
        desc_text.grid(row=row, column=1, columnspan=5, sticky="we", padx=6, pady=(6, 0))
        row += 1

        # Progress log
        ttk.Label(content_frame, text="Progress Log").grid(row=row, column=0, sticky="nw", pady=(10, 0))
        progress_frame = ttk.Frame(content_frame)
        progress_frame.grid(row=row, column=1, columnspan=5, sticky="we", padx=6, pady=(10, 0))
        progress_display = tk.Text(progress_frame, height=6, width=72, wrap="word")
        progress_display.insert("1.0", "")
        progress_display.config(state="disabled")
        progress_display.pack(fill=tk.BOTH, expand=False)
        ttk.Label(progress_frame, text="New progress entry:").pack(anchor="w", pady=(8, 2))
        new_progress_entry = tk.Text(progress_frame, height=4, width=72, wrap="word")
        new_progress_entry.pack(fill=tk.BOTH, expand=False)

        def _add_progress_entry_from_text(text_value):
            nonlocal staged_progress_entries
            text = text_value.strip()
            if not text:
                messagebox.showwarning("Progress", "Enter some progress text before adding.", parent=win)
                return
            now_str = date.today().isoformat()
            entry = f"[{now_str}] {text}\n"
            if task_id:
                try:
                    cur = self.db.conn.cursor()
                    cur.execute("SELECT progress_log FROM tasks WHERE id=?", (task_id,))
                    old = cur.fetchone()[0] or ""
                    new_log = entry + old
                    self.db.update_progress(task_id, new_log)
                    progress_display.config(state="normal")
                    progress_display.delete("1.0", tk.END)
                    progress_display.insert(tk.END, new_log)
                    progress_display.config(state="disabled")
                    new_progress_entry.delete("1.0", tk.END)
                    self._populate(); self._populate_kanban()
                except Exception:
                    logger.exception("Could not add progress")
                    messagebox.showerror("Progress Error", "Could not add progress", parent=win)
            else:
                staged_progress_entries = entry + staged_progress_entries
                progress_display.config(state="normal")
                progress_display.delete("1.0", tk.END)
                progress_display.insert(tk.END, staged_progress_entries)
                progress_display.config(state="disabled")
                new_progress_entry.delete("1.0", tk.END)

        ttk.Button(progress_frame, text="Add Progress Entry", command=lambda: _add_progress_entry_from_text(new_progress_entry.get("1.0", tk.END))).pack(pady=(6, 0))
        row += 1

        # Reminder Email (HTML) placed below progress section
        ttk.Label(content_frame, text="Reminder Email (HTML)").grid(row=row, column=0, sticky="nw", pady=(10, 0))
        email_body_text = tk.Text(content_frame, height=8, width=80, wrap="word")
        email_body_text.grid(row=row, column=1, columnspan=4, sticky="we", padx=6, pady=(10, 0))

        # Send Reminder Now button (next to email field)
        def _send_now_action():
            html_body = email_body_text.get("1.0", tk.END).strip()
            if not html_body:
                res = messagebox.askyesno("Send Empty Body?", "Reminder email body is empty. Send anyway?")
                if not res:
                    return
            # determine recipient
            label = responsible_var.get().strip()
            to_address = None
            if label and hasattr(responsible_cb, "lookup_map"):
                cid = responsible_cb.lookup_map.get(label)
                if cid:
                    cur = self.db.conn.cursor()
                    cur.execute("SELECT email FROM contacts WHERE id=?", (cid,))
                    rowc = cur.fetchone()
                    if rowc and rowc["email"]:
                        to_address = rowc["email"]
            if not to_address:
                # ask user to enter an email address
                to_address = simpledialog.askstring("Recipient", "Enter recipient email address:", parent=win)
                if not to_address:
                    messagebox.showinfo("Aborted", "No recipient provided; aborting send.", parent=win)
                    return
            # send via Outlook
            if not HAS_OUTLOOK:
                messagebox.showwarning("Outlook Unavailable", "Outlook integration is not available on this system.")
                return
            sent_ok = self._send_reminder_email(task_id or 0, to_address, title_var.get().strip() or "Task Reminder", html_body)
            if sent_ok:
                messagebox.showinfo("Sent", f"Reminder email sent to {to_address}.", parent=win)
            else:
                messagebox.showerror("Send Failed", "Failed to send reminder email (see logs).", parent=win)
                
        row += 1

        """# Centered Send Reminder Now button
        # Centered Send Reminder Now buttons (Outlook + Teams)
        btn_frame_send = ttk.Frame(content_frame)
        btn_frame_send.grid(row=row, column=0, columnspan=6, pady=(10, 0))
        # Outlook send (existing behavior)
        send_btn = ttk.Button(btn_frame_send, text="Send Reminder Now (Outlook)", command=_send_now_action)
        send_btn.pack(side=tk.LEFT, padx=(0,10))
        """
        # Teams send
        def _send_teams_now_action():
            html_body = email_body_text.get("1.0", tk.END).strip()
            if not html_body:
                res = messagebox.askyesno("Send Empty Body?", "Reminder body is empty. Send anyway?")
                if not res:
                    return
            # determine recipient the same way as for email
            label = responsible_var.get().strip()
            recipient = None
            if label and hasattr(responsible_cb, "lookup_map"):
                cid = responsible_cb.lookup_map.get(label)
                if cid:
                    cur = self.db.conn.cursor()
                    cur.execute("SELECT email FROM contacts WHERE id=?", (cid,))
                    rowc = cur.fetchone()
                    if rowc and rowc["email"]:
                        recipient = rowc["email"]
            if not recipient:
                # ask user to enter an email or webhook URL
                recipient = simpledialog.askstring("Recipient", "Enter recipient email address or Teams webhook URL:", parent=win)
                if not recipient:
                    messagebox.showinfo("Aborted", "No recipient provided; aborting Teams send.", parent=win)
                    return

            # format body for Teams: subject as bold heading then two blank lines then body
            subject = title_var.get().strip() or "Task Reminder"
            body = html_body  # plain text or HTML - webhook will display as markdown/text
            ok = self._send_teams_reminder(task_id or 0, recipient, subject, body)
            if ok:
                messagebox.showinfo("Teams", "Teams reminder delivered (or chat opened).", parent=win)
            else:
                messagebox.showerror("Teams", "Failed to deliver Teams reminder (see logs).", parent=win)
        # inside _open_edit_window where you have _send_now_action
                # Centered Send Reminder Now button
                    # Centered Send Reminder Now buttons (Outlook + Teams)
                    # --- Create the Send buttons (Outlook + Teams) ---
        # Make a small debug log so we can verify this code runs
                    # --- Create the Send buttons (Outlook + Teams) ---
        try:
            logger.debug("_open_edit_window: creating send buttons (Outlook + Teams)")
        except Exception:
            pass

        btn_frame_send = ttk.Frame(content_frame)
        btn_frame_send.grid(row=row, column=0, columnspan=6, pady=(10, 0))

        # Outlook send (existing behavior)
        try:
            send_btn = ttk.Button(btn_frame_send, text="Send Reminder Now (Outlook)", command=_send_now_action)
            send_btn.pack(side=tk.LEFT, padx=(0, 10))
        except Exception:
            logger.exception("Failed to create Outlook send button")

        # Teams send (use the _send_teams_now_action defined above)
        try:
            teams_btn = ttk.Button(btn_frame_send, text="Send Teams Reminder Now", command=_send_teams_now_action)
            teams_btn.pack(side=tk.LEFT, padx=(0, 10))
        except Exception:
            logger.exception("Failed to create Teams send button")

        # advance layout row index
        row += 1
        


        # .pack(side=tk.LEFT, padx=6)
        # Attachments
        ttk.Label(content_frame, text="Attachments").grid(row=row, column=0, sticky="nw", pady=(10, 0))
        attachments_frame = ttk.Frame(content_frame)
        attachments_frame.grid(row=row, column=1, columnspan=5, sticky="we", padx=6, pady=(10, 0))
        attachments_list_var = tk.StringVar(value=", ".join(os.path.basename(p) for p in existing_attachments))
        attachments_label = ttk.Label(attachments_frame, textvariable=attachments_list_var, wraplength=700)
        attachments_label.pack(anchor="w", fill=tk.X)

        def add_file_to_attachments():
            path = filedialog.askopenfilename(parent=win)
            if not path:
                return
            os.makedirs("attachments", exist_ok=True)
            fname = os.path.basename(path)
            dest = os.path.join("attachments", fname)
            if os.path.exists(dest):
                base, ext = os.path.splitext(fname)
                dest = os.path.join("attachments", f"{base}_{int(datetime.now().timestamp())}{ext}")
            with open(path, "rb") as fsrc, open(dest, "wb") as fdst:
                fdst.write(fsrc.read())
            if task_id:
                cur = self.db.conn.cursor()
                cur.execute("SELECT attachments FROM tasks WHERE id=?", (task_id,))
                rowa = cur.fetchone()
                files = []
                if rowa and rowa["attachments"]:
                    try:
                        files = json.loads(rowa["attachments"])
                    except Exception:
                        files = []
                files.append(dest)
                self.db.conn.execute("UPDATE tasks SET attachments=? WHERE id=?", (json.dumps(files), task_id))
                self.db.conn.commit()
                attachments_list_var.set(", ".join(os.path.basename(p) for p in files))
            else:
                staged_attachments.append(dest)
                attachments_list_var.set(", ".join(os.path.basename(p) for p in staged_attachments))

        def open_attachments():
            files = list(existing_attachments) + list(staged_attachments)
            if not files:
                messagebox.showinfo("Attachments", "No attachments to open.", parent=win)
                return
            for f in files:
                try:
                    if os.name == "nt":
                        os.startfile(f)
                    elif sys.platform == "darwin":
                        subprocess.run(["open", f])
                    else:
                        subprocess.run(["xdg-open", f])
                except Exception:
                    logger.exception("Could not open attachment")

        btns_attach = ttk.Frame(attachments_frame)
        btns_attach.pack(anchor="w", pady=(6, 0))
        ttk.Button(btns_attach, text="Add File", command=add_file_to_attachments).pack(side=tk.LEFT)
        ttk.Button(btns_attach, text="Open", command=open_attachments).pack(side=tk.LEFT, padx=6)
        row += 1

        # If editing existing task, load values now
        if task_id:
            try:
                cur = self.db.conn.cursor()
                cur.execute("SELECT * FROM tasks WHERE id=?", (task_id,))
                r = cur.fetchone()
                if r:
                    title_var.set(r["title"])
                    due_var.set(r["due_date"] or "")
                    priority_var.set(r["priority"] or "Medium")
                    status_var.set(r["status"] or "Pending")
                    desc_text.insert(tk.END, r["description"] or "")
                    if r["attachments"]:
                        try:
                            existing_attachments = json.loads(r["attachments"])
                        except Exception:
                            existing_attachments = []
                        attachments_list_var.set(", ".join(os.path.basename(p) for p in existing_attachments))
                    if r["reminder_minutes"]:
                        reminder_var.set(str(r["reminder_minutes"]))
                    if r["progress_log"]:
                        progress_display.config(state="normal")
                        progress_display.delete("1.0", tk.END)
                        progress_display.insert(tk.END, r["progress_log"])
                        progress_display.config(state="disabled")
                    rec_val = (r["recurrence"] or "").strip().lower()
                    if rec_val and rec_val != "none":
                        parsed = self._parse_recurrence(rec_val)
                        typ = parsed.get("type")
                        n = parsed.get("n", 1)
                        if typ == "days":
                            rec_type_var.set("Every N days")
                        elif typ == "weeks":
                            rec_type_var.set("Every N weeks")
                        elif typ == "months":
                            rec_type_var.set("Every N months")
                        else:
                            rec_type_var.set("None")
                        rec_n_var.set(str(n))
                    else:
                        rec_type_var.set("None")
                        rec_n_var.set("1")

                    # responsible
                    try:
                        repid = r["responsible_id"]
                        if repid:
                            label = self.db.get_contact_label(repid)
                            if label:
                                responsible_var.set(label)
                    except Exception:
                        pass

                    # reminder email body
                    try:
                        if r["reminder_email_body"]:
                            email_body_text.delete("1.0", tk.END)
                            email_body_text.insert(tk.END, r["reminder_email_body"])
                    except Exception:
                        pass
            except Exception:
                logger.exception("Error populating edit window")

        # Save & Cancel
        bottom_frame = ttk.Frame(win, padding=8)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X)
        sep = ttk.Separator(bottom_frame, orient="horizontal")
        sep.pack(fill=tk.X, pady=(0, 6))

        def _close():
            try:
                win.destroy()
            except Exception:
                pass

        def _save():
            title = title_var.get().strip()
            if not title:
                messagebox.showwarning("Validation", "Title is required", parent=win)
                return
            due = due_var.get().strip()
            if due:
                try:
                    datetime.strptime(due, "%Y-%m-%d")
                except ValueError:
                    messagebox.showwarning("Validation", "Date must be YYYY-MM-DD", parent=win)
                    return
            desc = desc_text.get("1.0", tk.END).strip()
            reminder_value = reminder_var.get().strip() or None
            if reminder_value not in (None, "", "None"):
                try:
                    reminder_minutes_int = int(reminder_value)
                except Exception:
                    messagebox.showwarning("Validation", "Reminder must be an integer number of minutes or blank.", parent=win)
                    return
                reminder_set_at_iso = datetime.now().isoformat(timespec="seconds")
            else:
                reminder_minutes_int = None
                reminder_set_at_iso = None

            rec_type_label = rec_type_var.get()
            try:
                n_val = int(rec_n_var.get())
                n_val = max(1, n_val)
            except Exception:
                messagebox.showwarning("Validation", "Recurrence interval N must be an integer >= 1.", parent=win)
                return
            if rec_type_label == "Every N days":
                rec_store = f"days:{n_val}"
            elif rec_type_label == "Every N weeks":
                rec_store = f"weeks:{n_val}"
            elif rec_type_label == "Every N months":
                rec_store = f"months:{n_val}"
            else:
                rec_store = "none"

            # determine responsible id
            responsible_label = responsible_var.get().strip()
            responsible_id_val = None
            if responsible_label and hasattr(responsible_cb, "lookup_map"):
                responsible_id_val = responsible_cb.lookup_map.get(responsible_label)

            reminder_email_html = email_body_text.get("1.0", tk.END).strip() or None
            try:
                if task_id:
                    self.db.update(task_id, title, desc, due or None, priority_var.get(), status_var.get(),
                                   reminder_minutes=reminder_minutes_int, reminder_set_at=reminder_set_at_iso, recurrence=rec_store,
                                   responsible_id=responsible_id_val, reminder_email_body=reminder_email_html)
                    if staged_attachments:
                        cur = self.db.conn.cursor()
                        cur.execute("SELECT attachments FROM tasks WHERE id=?", (task_id,))
                        rowa = cur.fetchone()
                        files = []
                        if rowa and rowa["attachments"]:
                            try:
                                files = json.loads(rowa["attachments"])
                            except Exception:
                                files = []
                        files.extend(staged_attachments)
                        self.db.conn.execute("UPDATE tasks SET attachments=? WHERE id=?", (json.dumps(files), task_id))
                else:
                    self.db.add(title, desc, due or None, priority_var.get(), status_var.get(),
                                reminder_minutes=reminder_minutes_int, reminder_set_at=reminder_set_at_iso, recurrence=rec_store,
                                responsible_id=responsible_id_val, reminder_email_body=reminder_email_html)
                    cur = self.db.conn.cursor()
                    cur.execute("SELECT last_insert_rowid() as id")
                    new_id = cur.fetchone()["id"]
                    if staged_attachments:
                        self.db.conn.execute("UPDATE tasks SET attachments=? WHERE id=?", (json.dumps(staged_attachments), new_id))
                    if staged_progress_entries:
                        self.db.update_progress(new_id, staged_progress_entries)
            except Exception:
                logger.exception("Save Error")
                messagebox.showerror("Save Error", "Could not save task", parent=win)
                return

            self._populate()
            self._populate_kanban()
            _close()

        btn_frame = ttk.Frame(bottom_frame)
        btn_frame.pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text="Save", command=_save).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(btn_frame, text="Cancel", command=_close).pack(side=tk.RIGHT, padx=(0, 6))

        try:
            win.focus_force()
            for child in content_frame.winfo_children():
                if isinstance(child, ttk.Entry):
                    child.focus_set()
                    break
        except Exception:
            pass

        def _on_win_configure(event=None):
            _on_frame_configure()
        win.bind("<Configure>", _on_win_configure)


    ###
    def _http_request_log(method, url, headers=None, json_payload=None, params=None):
        """
        Wrapper around requests.request with detailed logging of request/response.
        """
        try:
            logger.debug("HTTP %s %s headers=%s params=%s", method, url, headers, params)
            if json_payload:
                logger.debug("Payload: %s", str(json_payload)[:1000])  # limit length
            resp = requests.request(method, url, headers=headers, json=json_payload, params=params, timeout=30)
            logger.debug("Response status=%s body=%s", resp.status_code, resp.text[:2000])
            return resp
        except Exception:
            logger.exception("HTTP %s %s failed", method, url)
            raise
    # Add this method to TaskApp
    def _send_teams_reminder(self, task_id, to_address, subject_title, html_body):
        """
        Send Teams reminder. Try Graph/MSAL first (if configured). If that fails,
        fall back to browser automation with Selenium.

        This version attempts:
        1) Use real Chrome user-data-dir (so already-logged-in Teams web session is used).
        2) If Chrome fails to start (DevToolsActivePort / crash), retry with a temporary
            user-data-dir (clean profile) so ChromeLauncher won't collide with an existing Chrome.
        """
        import os
        import time
        import tempfile
        import shutil
        import logging
        import urllib.parse
        import re

        # setup file logging to app folder
        try:
            app_dir = os.path.dirname(os.path.abspath(__file__))
        except Exception:
            app_dir = os.getcwd()
        log_path = os.path.join(app_dir, "oats_teams.log")
        file_handler = None
        try:
            root_logger = logging.getLogger()
            # ensure single handler for log_path
            existing = [h for h in root_logger.handlers if getattr(h, "baseFilename", "") == log_path]
            if not existing:
                file_handler = logging.FileHandler(log_path, encoding="utf-8")
                file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
                root_logger.addHandler(file_handler)
                root_logger.setLevel(logging.DEBUG)
        except Exception:
            logger.exception("Could not install file logger")

        logger.debug("Attempt send Teams reminder to %s (task=%s)", to_address, task_id)

        # --- 1) Try Graph/MSAL (same approach you had) ---
        try:
            cfg = self.settings.get("graph", {})
            client_id = cfg.get("client_id")
            tenant = cfg.get("tenant_id") or cfg.get("tenant")
            scopes = cfg.get("scopes") or ["User.Read", "Chat.ReadWrite"]

            if client_id and tenant:
                logger.debug("Graph config present, attempting MSAL Graph path")
                try:
                    import msal, requests
                    authority = f"https://login.microsoftonline.com/{tenant}"
                    cache = msal.SerializableTokenCache()
                    token_cache_file = os.path.join(os.path.expanduser("~"), ".oats_graph_token_cache.bin")
                    if os.path.exists(token_cache_file):
                        try:
                            cache.deserialize(open(token_cache_file, "r").read())
                        except Exception:
                            pass
                    app = msal.PublicClientApplication(client_id=client_id, authority=authority, token_cache=cache)
                    accounts = app.get_accounts()
                    result = None
                    if accounts:
                        try:
                            result = app.acquire_token_silent(scopes, account=accounts[0])
                        except Exception:
                            result = None
                    if not result:
                        flow = app.initiate_device_flow(scopes=scopes)
                        if "user_code" not in flow:
                            logger.error("MSAL device flow initiation failed: %s", flow)
                            result = None
                        else:
                            messagebox.showinfo("Sign in required", f"To send Teams messages, sign in at {flow['verification_uri']} and enter the code {flow['user_code']}")
                            result = app.acquire_token_by_device_flow(flow)
                            try:
                                with open(token_cache_file, "w") as fhc:
                                    fhc.write(cache.serialize())
                            except Exception:
                                pass

                    if result and "access_token" in result:
                        access_token = result["access_token"]
                        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json"}

                        # Resolve user by email
                        user_get = requests.get(f"https://graph.microsoft.com/v1.0/users/{requests.utils.requote_uri(to_address)}", headers=headers)
                        if user_get.status_code == 404:
                            logger.error("Graph: Teams user not found for email %s", to_address)
                        else:
                            user_get.raise_for_status()
                            recipient_id = user_get.json().get("id")
                            if recipient_id:
                                me_r = requests.get("https://graph.microsoft.com/v1.0/me", headers=headers); me_r.raise_for_status()
                                me_id = me_r.json().get("id")
                                create_chat_payload = {
                                    "chatType": "oneOnOne",
                                    "members": [
                                        {
                                            "@odata.type": "#microsoft.graph.aadUserConversationMember",
                                            "roles": ["owner"],
                                            "user@odata.bind": f"https://graph.microsoft.com/v1.0/users('{me_id}')"
                                        },
                                        {
                                            "@odata.type": "#microsoft.graph.aadUserConversationMember",
                                            "roles": ["owner"],
                                            "user@odata.bind": f"https://graph.microsoft.com/v1.0/users('{recipient_id}')"
                                        }
                                    ]
                                }
                                create_r = requests.post("https://graph.microsoft.com/v1.0/chats", headers=headers, json=create_chat_payload)
                                chat_id = None
                                if create_r.status_code in (200, 201):
                                    chat_id = create_r.json().get("id")
                                else:
                                    logger.debug("Graph: create chat responded %s; trying to locate existing chat", create_r.status_code)
                                    list_chats = requests.get("https://graph.microsoft.com/v1.0/me/chats", headers=headers, params={"$top": 50})
                                    list_chats.raise_for_status()
                                    for c in list_chats.json().get("value", []):
                                        try:
                                            members_r = requests.get(f"https://graph.microsoft.com/v1.0/chats/{c['id']}/members", headers=headers); members_r.raise_for_status()
                                            for m in members_r.json().get("value", []):
                                                if m.get("userId") == recipient_id:
                                                    chat_id = c["id"]; break
                                            if chat_id: break
                                        except Exception:
                                            continue
                                if chat_id:
                                    message_html = f"<b>{(subject_title or 'Reminder').strip()}</b><br><br>{html_body or ''}"
                                    send_payload = {"body": {"contentType": "html", "content": message_html}}
                                    send_r = requests.post(f"https://graph.microsoft.com/v1.0/chats/{chat_id}/messages", headers=headers, json=send_payload)
                                    if send_r.status_code in (200, 201):
                                        logger.info("Graph: Teams message sent to %s (task=%s)", to_address, task_id)
                                        return True
                                    else:
                                        logger.error("Graph send failed: %s %s", send_r.status_code, send_r.text)
                    else:
                        logger.debug("Graph: token not acquired or login aborted; will fallback to browser UI.")
                except Exception:
                    logger.exception("Graph path raised exception; falling back to browser automation")
            else:
                logger.debug("No Graph client_id/tenant present in settings; skip Graph")
        except Exception:
            logger.exception("Unhandled error during Graph/MSAL attempt (ignored)")

        # --- 2) Browser fallback using Selenium ---
        logger.debug("Falling back to browser automation for Teams web.")

        try:
            from selenium import webdriver
            from selenium.common.exceptions import SessionNotCreatedException, WebDriverException
            from selenium.webdriver.common.by import By
            from selenium.webdriver.common.keys import Keys
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            from webdriver_manager.chrome import ChromeDriverManager
        except Exception:
            logger.exception("Selenium/webdriver-manager not installed - browser fallback unavailable")
            return False

        # build chat url for direct chat to email
        users_q = urllib.parse.quote(to_address)
        chat_url = f"https://teams.microsoft.com/l/chat/0/0?users={users_q}"

        def _start_driver_with_options(options, driver_path=None, timeout=60):
            """Start chrome driver using webdriver-manager (or provided driver_path). Return driver or raise."""
            try:
                # If driver_path provided use it (packaged driver); else use webdriver-manager
                if driver_path:
                    service = webdriver.chrome.service.Service(driver_path)
                else:
                    chromedriver_path = ChromeDriverManager().install()
                    service = webdriver.chrome.service.Service(chromedriver_path)
                driver = webdriver.Chrome(service=service, options=options)
                driver.set_page_load_timeout(timeout)
                return driver
            except Exception:
                raise

        # prepare options common
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-infobars")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        # don't use headless - want interactive session (and Teams web may block headless)
        # options.add_argument("--headless=new")  # not recommended for Teams

        # Try user's Chrome profile first (best UX)
        tried_temp_profile = False
        temp_profile_dir = None
        driver = None
        try_user_profile = True
        user_profile_dir = None
        try:
            if os.name == "nt":
                user_profile_dir = os.path.join(os.environ.get("USERPROFILE", ""), "AppData", "Local", "Google", "Chrome", "User Data")
            elif sys.platform == "darwin":
                user_profile_dir = os.path.expanduser("~/Library/Application Support/Google/Chrome")
            else:
                user_profile_dir = os.path.expanduser("~/.config/google-chrome")

            if user_profile_dir and os.path.exists(user_profile_dir):
                try:
                    options.add_argument(f"--user-data-dir={user_profile_dir}")
                    logger.debug("Using Chrome user-data-dir: %s", user_profile_dir)
                except Exception:
                    logger.exception("Could not set --user-data-dir (ignoring)")
            else:
                try_user_profile = False
                logger.debug("No existing Chrome user-data-dir found, skipping user profile attempt")
        except Exception:
            logger.exception("Error while determining user profile dir")

        # Attempt loop: try user profile -> on failure remove that flag and try temp profile
        last_exception = None
        for attempt in range(2):
            try:
                logger.debug("Chrome start attempt #%s (using user profile? %s)", attempt + 1, try_user_profile)
                driver = _start_driver_with_options(options)
                logger.debug("Chrome started successfully (attempt %s)", attempt + 1)
                break
            except Exception as e:
                last_exception = e
                logger.exception("Chrome start failed on attempt %s", attempt + 1)
                # if first attempt used real profile, retry with temp profile
                if try_user_profile:
                    logger.debug("Will retry with a temporary profile to avoid profile lock/crash")
                    # remove user-data-dir arg and create a temp dir
                    try:
                        # remove any existing --user-data-dir from options
                        new_args = [a for a in options.arguments if not a.startswith("--user-data-dir")]
                        # unfortunately ChromeOptions doesn't expose simple removal API; recreate options
                        options = webdriver.ChromeOptions()
                        options.add_argument("--disable-infobars"); options.add_argument("--no-sandbox")
                        options.add_argument("--disable-dev-shm-usage"); options.add_argument("--disable-gpu")
                        temp_profile_dir = tempfile.mkdtemp(prefix="oats_chrome_profile_")
                        options.add_argument(f"--user-data-dir={temp_profile_dir}")
                        logger.debug("Created temporary chrome profile at %s", temp_profile_dir)
                        tried_temp_profile = True
                        tried_user_profile = False
                        try_user_profile = False  # ensure next loop uses temp profile
                        continue  # next attempt
                    except Exception:
                        logger.exception("Failed to prepare temporary profile; aborting browser fallback")
                        break
                else:
                    # already tried temp profile or nothing left to try
                    break

        if driver is None:
            logger.error("Failed to start Chrome for browser automation. Last error: %s", repr(last_exception))
            # cleanup temp profile if created
            if temp_profile_dir:
                try:
                    shutil.rmtree(temp_profile_dir, ignore_errors=True)
                except Exception:
                    pass
            return False

        # driver is running; navigate and attempt to send
        try:
            logger.debug("Navigating to Teams chat URL: %s", chat_url)
            driver.get(chat_url)
            wait = WebDriverWait(driver, 60)
            # wait for a contenteditable input area
            try:
                editable = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "[contenteditable='true']")))
            except Exception:
                logger.debug("Editable area not found quickly; waiting up to 120s (login or SSO possible)")
                try:
                    editable = WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.CSS_SELECTOR, "[contenteditable='true']")))
                except Exception:
                    logger.error("Could not find Teams chat input area; leaving browser open for manual send")
                    return False

            # Compose plain text fallback from HTML
            text_body = re.sub(r"<[^>]+>", "", html_body or "")
            subject = (subject_title or "Reminder").strip()
            to_send = f"{subject}\n\n{text_body}"

            # click and send
            try:
                editable.click()
            except Exception:
                pass
            # send message (typing then Enter)
            editable.send_keys(to_send)
            time.sleep(0.4)
            editable.send_keys(Keys.ENTER)
            logger.info("Browser automation: Teams message sent/queued to %s (task=%s)", to_address, task_id)
            time.sleep(1.2)
            try:
                driver.quit()
            except Exception:
                pass
            # cleanup temp profile if created
            if temp_profile_dir:
                try:
                    shutil.rmtree(temp_profile_dir, ignore_errors=True)
                except Exception:
                    pass
            return True
        except Exception:
            logger.exception("Error while automating Teams web to send message")
            try:
                driver.quit()
            except Exception:
                pass
            if temp_profile_dir:
                try:
                    shutil.rmtree(temp_profile_dir, ignore_errors=True)
                except Exception:
                    pass
            return False
        finally:
            try:
                if file_handler is not None:
                    root_logger.removeHandler(file_handler)
                    file_handler.close()
            except Exception:
                pass
    # -------------------- Other CRUD helpers & Kanban --------------------

        # Recurrence helpers
    def _parse_recurrence(self, rec_str):
        """
        Parse stored recurrence strings like 'days:3', 'weeks:1', 'months:2' or 'none'.
        Returns dict: {'type': 'days'|'weeks'|'months'|'none', 'n': int}
        """
        try:
            if not rec_str:
                return {"type": "none", "n": 0}
            s = str(rec_str).strip().lower()
            if s in ("none", ""):
                return {"type": "none", "n": 0}
            if ":" in s:
                typ, n = s.split(":", 1)
                typ = typ.strip()
                try:
                    n = max(1, int(n))
                except Exception:
                    n = 1
                if typ in ("days", "weeks", "months"):
                    return {"type": typ, "n": n}
            # fallback
            return {"type": "none", "n": 0}
        except Exception:
            return {"type": "none", "n": 0}

    def _compute_next_due(self, due_date_iso, recurrence_store):
        """
        Given a due date string 'YYYY-MM-DD' and a recurrence store like 'days:3',
        compute next due date string in same format. Returns None on failure.
        """
        try:
            if not due_date_iso:
                return None
            parsed = self._parse_recurrence(recurrence_store)
            typ = parsed.get("type")
            n = parsed.get("n", 0) or 0
            if typ == "none" or n <= 0:
                return None
            cur_due = datetime.strptime(due_date_iso, "%Y-%m-%d").date()
            if typ == "days":
                next_due = cur_due + timedelta(days=n)
            elif typ == "weeks":
                next_due = cur_due + timedelta(weeks=n)
            elif typ == "months":
                # Add months conservatively (roll month forward, clamp day)
                year = cur_due.year
                month = cur_due.month + n
                # normalize year/month
                year += (month - 1) // 12
                month = ((month - 1) % 12) + 1
                day = min(cur_due.day, calendar.monthrange(year, month)[1])
                next_due = date(year, month, day)
            else:
                return None
            return next_due.isoformat()
        except Exception:
            logger.exception("Error computing next due date")
            return None
        
    def _selected_tree_task_id(self):
        sel = self.tree.selection()
        if not sel:
            return None
        try:
            return int(self.tree.item(sel[0], "values")[0])
        except Exception:
            return None

    def _open_selected_kanban_attachments(self):
        if not self.kanban_selected_id:
            messagebox.showwarning("No Task", "Please select a task first.")
            return
        cur = self.db.conn.cursor()
        cur.execute("SELECT attachments FROM tasks WHERE id=?", (self.kanban_selected_id,))
        row = cur.fetchone()
        if not row or not row["attachments"]:
            messagebox.showinfo("No Attachments", "No attachments found for this task.")
            return
        files = json.loads(row["attachments"])
        for f in files:
            try:
                if os.name == "nt":
                    os.startfile(f)
                elif sys.platform == "darwin":
                    subprocess.run(["open", f])
                else:
                    subprocess.run(["xdg-open", f])
            except Exception:
                logger.exception("Could not open attachment")

    def _add_attachment(self):
        path = filedialog.askopenfilename()
        if not path:
            return
        os.makedirs("attachments", exist_ok=True)
        fname = os.path.basename(path)
        dest = os.path.join("attachments", fname)
        if os.path.exists(dest):
            base, ext = os.path.splitext(fname)
            dest = os.path.join("attachments", f"{base}_{int(datetime.now().timestamp())}{ext}")
        with open(path, "rb") as fsrc, open(dest, "wb") as fdst:
            fdst.write(fsrc.read())
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("No Task", "Select a task first.")
            return
        task_id = int(self.tree.item(sel[0], "values")[0])
        cur = self.db.conn.cursor()
        cur.execute("SELECT attachments FROM tasks WHERE id=?", (task_id,))
        row = cur.fetchone()
        files = []
        if row and row["attachments"]:
            try:
                files = json.loads(row["attachments"])
            except Exception:
                files = []
        files.append(dest)
        self.db.conn.execute("UPDATE tasks SET attachments=? WHERE id=?", (json.dumps(files), task_id))
        self.db.conn.commit()
        self.attachments_var.set(", ".join(os.path.basename(f) for f in files))
        messagebox.showinfo("Attachment", f"File {os.path.basename(dest)} added.")

    def _open_attachment(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("No Task", "Select a task first.")
            return
        task_id = int(self.tree.item(sel[0], "values")[0])
        cur = self.db.conn.cursor()
        cur.execute("SELECT attachments FROM tasks WHERE id=?", (task_id,))
        row = cur.fetchone()
        if not row or not row["attachments"]:
            messagebox.showinfo("No Attachments", "No attachments found for this task.")
            return
        files = json.loads(row["attachments"])
        for f in files:
            try:
                if os.name == "nt":
                    os.startfile(f)
                elif sys.platform == "darwin":
                    subprocess.run(["open", f])
                else:
                    subprocess.run(["xdg-open", f])
            except Exception:
                logger.exception("Could not open attachment")

    def _on_kanban_drag_start(self, event):
        lb = event.widget
        idx = lb.nearest(event.y)
        if idx >= 0:
            self.drag_data = {"listbox": lb, "index": idx, "task_line": lb.get(idx)}

    ##
    # add this method to TaskApp (anywhere inside the class)
    def _global_kanban_double_click(self, event):
        try:
            logger.debug("global handler invoked: widget=%s coords=(%s,%s) type=%s",
                        getattr(event, "widget", None),
                        getattr(event, "x_root", None),
                        getattr(event, "y_root", None),
                        getattr(event, "type", None))

            x = getattr(event, "x_root", None)
            y = getattr(event, "y_root", None)

            widget = None
            if x is not None and y is not None:
                widget = self.winfo_containing(x, y)
                logger.debug("winfo_containing -> %s", widget)

            if not widget:
                widget = getattr(event, "widget", None)
                logger.debug("using event.widget -> %s", widget)

            # If we hit a Canvas, try to resolve an embedded window/frame at the canvas coords
            if isinstance(widget, tk.Canvas):
                try:
                    # convert screen coords to canvas coords
                    cx = widget.canvasx(x - widget.winfo_rootx())
                    cy = widget.canvasy(y - widget.winfo_rooty())
                    items = widget.find_overlapping(cx, cy, cx, cy)
                    for it in items:
                        try:
                            win_name = widget.itemcget(it, "window")
                            if win_name:
                                try:
                                    child = widget.nametowidget(win_name)
                                    if child:
                                        widget = child
                                        logger.debug("Resolved embedded child widget %s from canvas item %s", child, it)
                                        break
                                except Exception:
                                    pass
                        except Exception:
                            pass
                except Exception:
                    pass

            if not widget:
                logger.debug("no widget under pointer; aborting")
                return

            # climb parents and check for our custom attribute _kanban_task_id
            cur = widget
            chain = []
            while cur:
                chain.append(str(cur))
                try:
                    tid = getattr(cur, "_kanban_task_id", None)
                    logger.debug("checking widget %s for _kanban_task_id -> %s", cur, tid)
                    if tid:
                        try:
                            tid_int = int(tid)
                        except Exception:
                            tid_int = None
                        if tid_int:
                            logger.debug("Found kanban task id=%s on widget %s — opening editor", tid_int, cur)
                            try:
                                self._open_edit_window(tid_int)
                            except Exception:
                                logger.exception("Error opening editor from global double-click")
                            return
                except Exception:
                    logger.exception("Error inspecting widget for kanban id")
                try:
                    cur = getattr(cur, "master", None)
                except Exception:
                    break

            logger.debug("No kanban task id found in parent chain. widget chain: %s", " -> ".join(chain))
        except Exception:
            logger.exception("Unhandled error in global kanban double-click handler")


    ##
    def _on_kanban_double_click(self, event, lb=None):
        try:
            widget = lb if lb is not None else event.widget
            idx = widget.nearest(event.y)
            if idx < 0:
                return
            status = getattr(widget, "status_name", None) or "Pending"
            try:
                task_id = self.kanban_item_map[status][idx]
            except Exception:
                # fallback: attempt to find by title
                try:
                    title = widget.get(idx)
                    cur = self.db.conn.cursor()
                    cur.execute("SELECT id FROM tasks WHERE title=? LIMIT 1", (title,))
                    row = cur.fetchone()
                    if row:
                        task_id = row["id"]
                    else:
                        return
                except Exception:
                    return
            # open the editor
            try:
                self._open_edit_window(int(task_id))
            except Exception:
                logger.exception("Error opening task editor from kanban double-click")
        except Exception:
            logger.exception("Unhandled error in _on_kanban_double_click")


    def _on_kanban_drag_motion(self, event):
        lb = event.widget
        lb.selection_clear(0, tk.END)
        lb.selection_set(lb.nearest(event.y))

    def _on_kanban_drag_drop(self, event):
        if not hasattr(self, "drag_data"):
            return
        src_lb = self.drag_data["listbox"]
        src_status = src_lb.status_name
        src_index = self.drag_data["index"]
        try:
            task_id = self.kanban_item_map[src_status][src_index]
        except Exception:
            self.drag_data = None
            return
        src_lb.delete(src_index)
        try:
            del self.kanban_item_map[src_status][src_index]
        except Exception:
            pass
        widget = event.widget.winfo_containing(event.x_root, event.y_root)
        target_lb = widget if isinstance(widget, tk.Listbox) else None
        if target_lb and hasattr(target_lb, "status_name"):
            target_status = target_lb.status_name
            try:
                cur = self.db.conn.cursor()
                cur.execute("SELECT title FROM tasks WHERE id=?", (task_id,))
                row = cur.fetchone()
                display = row["title"] if row else "Untitled"
                target_lb.insert(tk.END, display)
            except Exception:
                try:
                    target_lb.insert(tk.END, "Untitled")
                except Exception:
                    pass
            self.kanban_item_map[target_status].append(task_id)
            self._move_task(task_id, target_status)
        else:
            try:
                cur = self.db.conn.cursor()
                row = cur.execute("SELECT title FROM tasks WHERE id=?", (task_id,)).fetchone()
                title = row["title"] if row else "Untitled"
                src_lb.insert(tk.END, title)
                self.kanban_item_map[src_status].append(task_id)
            except Exception:
                pass
        self.drag_data = None

    def _open_task_on_doubleclick(self, event, widget):
        try:
            task_id = None
            cls_name = widget.winfo_class()
            if cls_name == 'Treeview' or hasattr(widget, 'item'):
                item = widget.focus() or widget.identify_row(event.y)
                if not item:
                    return
                vals = widget.item(item).get('values', ())
                if vals:
                    task_id = vals[0]
                else:
                    task_id = item
            elif cls_name == 'Listbox' or hasattr(widget, 'curselection'):
                sel = widget.curselection()
                if not sel:
                    return
                idx = sel[0]
                try:
                    task_text = widget.get(idx)
                    parts = str(task_text).split('-', 1)
                    task_id = parts[0].strip()
                except Exception:
                    task_id = idx
            else:
                return
            if task_id is None:
                return
            try:
                tid = int(task_id)
                self._open_edit_window(tid)
                return
            except Exception:
                pass
            edit_func_names = ('open_edit_window', 'open_edit_task_window', 'open_task_editor', 'edit_task', 'edit_task_window', 'open_task')
            for name in edit_func_names:
                fn = globals().get(name)
                if callable(fn):
                    try:
                        fn(task_id)
                        return
                    except Exception:
                        continue
            messagebox.showinfo("Edit task", f"Double-clicked task: {task_id}\nCould not open editor.")
        except Exception:
            logger.exception("Unexpected error in double-click handler")

    # -------------------- Save / Delete / Mark Done --------------------
    def _save_inline_from_form(self):
        sel = self.tree.selection()
        if sel:
            try:
                task_id = int(self.tree.item(sel[0], "values")[0])
            except Exception:
                messagebox.showwarning("Save", "Could not determine selected task id.")
                return
            self._open_edit_window(task_id)
        else:
            self._open_edit_window(None)

    def _on_task_double_click(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            task_id = int(self.tree.item(sel[0], "values")[0])
        except Exception:
            return
        self._open_edit_window(task_id)

    def _apply_filters(self):
        fd = self.filter_due_var.get().strip() if hasattr(self, "filter_due_var") else ""
        if fd:
            try:
                datetime.strptime(fd, "%Y-%m-%d")
            except ValueError:
                messagebox.showwarning("Filter", "Due Date filter must be YYYY-MM-DD")
                return
        # Refresh all views so filters are applied everywhere
        try:
            self._populate()
        except Exception:
            logger.exception("Error populating Task List from _apply_filters")
        try:
            self._populate_kanban()
        except Exception:
            logger.exception("Error populating Kanban from _apply_filters")
        try:
            self._populate_trash()
        except Exception:
            logger.exception("Error populating Trash from _apply_filters")

    def _clear_filters(self):
        if hasattr(self, "filter_text_var"):
            self.filter_text_var.set("")
        if hasattr(self, "filter_priority_var"):
            self.filter_priority_var.set("All")
        if hasattr(self, "filter_status_var"):
            self.filter_status_var.set("All")
        if hasattr(self, "filter_due_var"):
            self.filter_due_var.set("")
        self._populate()

    def _populate(self):
        try:
            for rowid in self.tree.get_children():
                self.tree.delete(rowid)
        except Exception:
            pass
        try:
            rows = self.db.fetch()
        except Exception:
            rows = []

        ft = (self.filter_text_var.get().strip().lower() if hasattr(self, "filter_text_var") else "").strip()
        fpri = (self.filter_priority_var.get() if hasattr(self, "filter_priority_var") else "All")
        fstat = (self.filter_status_var.get() if hasattr(self, "filter_status_var") else "All")
        fdue = (self.filter_due_var.get().strip() if hasattr(self, "filter_due_var") else "").strip()
        fshow_completed = (self.filter_show_completed_var.get() if hasattr(self, "filter_show_completed_var") else "Yes")

        def row_matches(r):
            if ft:
                hay = ((r["title"] or "") + " " + (r["description"] or "")).lower()
                if ft not in hay:
                    return False
            if fpri and fpri != "All":
                if (r["priority"] or "") != fpri:
                    return False
            if fstat and fstat != "All":
                if (r["status"] or "") != fstat:
                    return False
            if fdue:
                try:
                    if (r["due_date"] or "") != fdue:
                        return False
                except Exception:
                    return False
            return True

        insert_index = 0
        for r in rows:
            try:
                status_val = (r["status"] or "").strip()
                if fshow_completed == "No" and status_val.lower() == "done":
                    continue
                if not row_matches(r):
                    continue
                desc = r["description"] or ""
                desc_preview = desc.replace("<body>", "").replace("</body>", "").replace("<html>", "").replace("</html>", "")
                desc_preview = desc_preview.replace("\n", " ")
                if len(desc_preview) > 80:
                    desc_preview = desc_preview[:80] + "..."

                reminder_val = r["reminder_minutes"] if "reminder_minutes" in r.keys() else None
                reminder_display = str(reminder_val) if reminder_val not in (None, "", "None") else "—"

                title_display = r["title"] or ""
                desc_display = desc_preview

                is_done = status_val.lower() == "done"

                responsible_label = self.db.get_contact_label(r["responsible_id"]) if r["responsible_id"] else ""

                if self.settings.get("show_description", False):
                    values = [
                        r["id"],
                        title_display,
                        desc_display,
                        r["due_date"] or "—",
                        r["priority"],
                        r["status"],
                        responsible_label,
                        reminder_display
                    ]
                else:
                    values = [
                        r["id"],
                        title_display,
                        r["due_date"] or "—",
                        r["priority"],
                        r["status"],
                        responsible_label,
                        reminder_display
                    ]

                tags = []
                if is_done:
                    tags.append("completed")
                pr = (r["priority"] or "").lower()
                if pr == "high":
                    tags.append("priority_high")
                elif pr == "medium":
                    tags.append("priority_medium")
                else:
                    tags.append("priority_low")
                tags.append("evenrow" if insert_index % 2 == 0 else "oddrow")

                try:
                    iid = self.tree.insert("", tk.END, values=values, tags=tags)
                    if is_done:
                        try:
                            self.tree.tag_configure("completed", foreground="#666666")
                            if hasattr(self, "strike_font"):
                                self.tree.tag_configure("completed", font=self.strike_font)
                        except Exception:
                            pass
                except Exception:
                    try:
                        self.tree.insert("", tk.END, values=values)
                    except Exception:
                        pass

                insert_index += 1
            except Exception:
                logger.exception("Error inserting row in _populate")
                continue


    ####
    def _kanban_click_select(self, event, lb):
        """
        Handle single-click on a Kanban listbox item: select and populate details,
        but do NOT open the edit window. Double-click will open the editor.
        """
        try:
            # find clicked index
            idx = lb.nearest(event.y)
            if idx < 0:
                return

            # set listbox selection (visual)
            lb.selection_clear(0, tk.END)
            lb.selection_set(idx)
            lb.activate(idx)

            status = getattr(lb, "status_name", None) or "Pending"
            try:
                task_id = self.kanban_item_map[status][idx]
            except Exception:
                # nothing to select
                return

            # set selected id/status so other actions/buttons use it
            self.kanban_selected_id = task_id
            self.kanban_selected_status = status

            # populate details panel (reuse existing logic)
            cur = self.db.conn.cursor()
            cur.execute("SELECT description, progress_log, outlook_id, attachments FROM tasks WHERE id=?", (task_id,))
            row = cur.fetchone()
            if not row:
                return
            desc = row["description"] or ""
            prog = row["progress_log"] or ""
            outlook_id = row["outlook_id"]

            # show HTML if available else plain text (reuse your code path)
            if outlook_id and HAS_HTML:
                clean = desc or ""
                clean = re.sub(r'<style.*?>.*?</style>', '', clean, flags=re.DOTALL | re.IGNORECASE)
                clean = re.sub(r'<font[^>]*>', '', clean, flags=re.IGNORECASE).replace("</font>", "")
                clean = re.sub(r'style="[^"]*font-size:[^";]*;?"', '', clean, flags=re.IGNORECASE)
                clean = re.sub(r'style="[^"]*font-family:[^";]*;?"', '', clean, flags=re.IGNORECASE)
                clean = re.sub(r'<span[^>]*>', '<span>', clean, flags=re.IGNORECASE)
                if os.name == "nt":
                    wrapper_style = "font-family:Segoe UI, Arial; font-size:9pt; line-height:1.3; color:#333;"
                else:
                    wrapper_style = "font-family:Arial; font-size:11px; line-height:1.3; color:#333;"
                clean = f"<div style='{wrapper_style}'>{clean}</div>"
                try:
                    # switch to HTML label if present
                    if hasattr(self, "kanban_text") and self.kanban_text is not None:
                        try:
                            self.kanban_text.pack_forget()
                        except Exception:
                            pass
                    self.kanban_html.set_html(clean)
                    self.kanban_html.pack(fill=tk.BOTH, expand=True)
                except Exception:
                    # fallback to plain text
                    if hasattr(self, "kanban_html"):
                        try:
                            self.kanban_html.pack_forget()
                        except Exception:
                            pass
                    self.kanban_text.delete("1.0", tk.END)
                    self.kanban_text.insert(tk.END, desc)
                    self.kanban_text.pack(fill=tk.BOTH, expand=True)
            else:
                try:
                    if HAS_HTML:
                        self.kanban_html.pack_forget()
                except Exception:
                    pass
                self.kanban_text.delete("1.0", tk.END)
                self.kanban_text.insert(tk.END, desc)
                self.kanban_text.pack(fill=tk.BOTH, expand=True)

            # progress
            self.kanban_progress.delete("1.0", tk.END)
            self.kanban_progress.insert(tk.END, prog)

            # attachments
            try:
                files = []
                if row["attachments"]:
                    files = json.loads(row["attachments"])
                self.kanban_attachments_var.set(", ".join(os.path.basename(f) for f in files) if files else "No attachments")
            except Exception:
                self.kanban_attachments_var.set("No attachments")

            # enable action buttons
            self.btn_edit.config(state="normal")
            self.btn_delete.config(state="normal")
            self.btn_done.config(state="normal")
            self.btn_prev.config(state="normal" if self.kanban_selected_status != "Pending" else "disabled")
            self.btn_next.config(state="normal" if self.kanban_selected_status != "Done" else "disabled")
        except Exception:
            logger.exception("Error in kanban single-click select")

    ###
    def _create_kanban_card(self, parent, task_row):
        """
        Create a 'card' in `parent` (an inner frame for the column).
        Single-click selects (populates details & enables buttons).
        Double-click opens editor.
        """

        # helper to safely read sqlite3.Row values with a default
        def _val(r, key, default=""):
            try:
                if key in r.keys() and r[key] is not None:
                    return r[key]
            except Exception:
                try:
                    # fallback: index access
                    v = r[key]
                    return v if v is not None else default
                except Exception:
                    return default
            return default

        try:
            tid = None
            try:
                tid = int(_val(task_row, "id", None))
            except Exception:
                tid = None

            title = str(_val(task_row, "title", "(no title)")).strip()
            due = str(_val(task_row, "due_date", "") or "")
            pr = str(_val(task_row, "priority", "Medium")).lower()
            responsible_label = ""
            try:
                resp_id = _val(task_row, "responsible_id", None)
                if resp_id:
                    responsible_label = self.db.get_contact_label(resp_id) or ""
            except Exception:
                responsible_label = ""

            # choose bg by priority
            if pr == "high":
                bg = "#FFD6D6"
            elif pr == "medium":
                bg = "#FFF5CC"
            else:
                bg = "#E6FFEA"

            # wrapper gives the colored card background
            wrapper = tk.Frame(parent, bg=bg, bd=1, relief="flat")
            wrapper.pack(fill=tk.X, pady=(6, 4), padx=6)

            # inner content frame to get padding
            content = tk.Frame(wrapper, bg=bg)
            content.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

            lbl_title = tk.Label(content, text=title, bg=bg, fg="#111111", anchor="w", justify="left",
                                font=("", 10, "bold"), wraplength=320)
            lbl_title.pack(fill=tk.X, anchor="w")

            meta_parts = [f"ID:{tid}" if tid is not None else "ID:?"]
            if due:
                meta_parts.append(f"Due: {due}")
            if responsible_label:
                meta_parts.append(responsible_label)
            lbl_meta = tk.Label(content, text="  •  ".join(meta_parts), bg=bg, fg="#333333", anchor="w",
                                justify="left", font=("", 9), wraplength=320)
            lbl_meta.pack(fill=tk.X, anchor="w", pady=(4, 0))

            lbl_priority = tk.Label(content, text=str(_val(task_row, "priority", "")), bg=bg, fg="#222222",
                                    anchor="w", justify="left", font=("", 8))
            lbl_priority.pack(anchor="w", pady=(6, 0))

            # register widget so we can un-highlight later
            try:
                if tid is not None:
                    self.kanban_card_widgets[tid] = wrapper
            except Exception:
                pass

            # Attach task id to wrapper and key child widgets so winfo_containing results can be resolved quickly
            try:
                wrapper._kanban_task_id = tid
                content._kanban_task_id = tid
                lbl_title._kanban_task_id = tid
                lbl_meta._kanban_task_id = tid
                lbl_priority._kanban_task_id = tid
            except Exception:
                pass

            # --- selection (single click) ---
            def _select_card(e=None, _id=tid, _widget=wrapper):
                try:
                    # visual highlight: remove previous
                    prev = getattr(self, "_kanban_highlighted", None)
                    if prev is not None and prev is not _widget:
                        try:
                            prev.configure(highlightthickness=0)
                        except Exception:
                            pass
                    # highlight current
                    try:
                        _widget.configure(highlightbackground="#3b82f6", highlightcolor="#3b82f6", highlightthickness=2)
                        self._kanban_highlighted = _widget
                    except Exception:
                        pass

                    # set selected id and status
                    self.kanban_selected_id = _id
                    self.kanban_selected_status = None
                    for s, idlist in self.kanban_item_map.items():
                        if _id in idlist:
                            self.kanban_selected_status = s
                            break

                    # populate details panel
                    cur = self.db.conn.cursor()
                    cur.execute("SELECT description, progress_log, outlook_id, attachments FROM tasks WHERE id=?", (_id,))
                    row = cur.fetchone()
                    if not row:
                        return
                    desc = row["description"] or ""
                    prog = row["progress_log"] or ""
                    # safe access for outlook_id (sqlite3.Row has no .get)
                    try:
                        outlook_id = row["outlook_id"] if ("outlook_id" in row.keys() and row["outlook_id"] is not None) else None
                    except Exception:
                        outlook_id = None

                    # show HTML if available else plain text (reuse your code path)
                    if outlook_id and HAS_HTML:
                        clean = desc or ""
                        clean = re.sub(r'<style.*?>.*?</style>', '', clean, flags=re.DOTALL | re.IGNORECASE)
                        clean = re.sub(r'<font[^>]*>', '', clean, flags=re.IGNORECASE).replace("</font>", "")
                        clean = re.sub(r'style="[^"]*font-size:[^";]*;?"', '', clean, flags=re.IGNORECASE)
                        clean = re.sub(r'style="[^"]*font-family:[^";]*;?"', '', clean, flags=re.IGNORECASE)
                        clean = re.sub(r'<span[^>]*>', '<span>', clean, flags=re.IGNORECASE)
                        if os.name == "nt":
                            wrapper_style = "font-family:Segoe UI, Arial; font-size:9pt; line-height:1.3; color:#333;"
                        else:
                            wrapper_style = "font-family:Arial; font-size:11px; line-height:1.3; color:#333;"
                        clean = f"<div style='{wrapper_style}'>{clean}</div>"
                        try:
                            if hasattr(self, "kanban_text") and self.kanban_text is not None:
                                try:
                                    self.kanban_text.pack_forget()
                                except Exception:
                                    pass
                            self.kanban_html.set_html(clean)
                            self.kanban_html.pack(fill=tk.BOTH, expand=True)
                        except Exception:
                            if hasattr(self, "kanban_html"):
                                try:
                                    self.kanban_html.pack_forget()
                                except Exception:
                                    pass
                            self.kanban_text.delete("1.0", tk.END)
                            self.kanban_text.insert(tk.END, desc)
                            self.kanban_text.pack(fill=tk.BOTH, expand=True)
                    else:
                        try:
                            if HAS_HTML:
                                self.kanban_html.pack_forget()
                        except Exception:
                            pass
                        self.kanban_text.delete("1.0", tk.END)
                        self.kanban_text.insert(tk.END, desc)
                        self.kanban_text.pack(fill=tk.BOTH, expand=True)

                    # progress & attachments
                    self.kanban_progress.delete("1.0", tk.END)
                    self.kanban_progress.insert(tk.END, prog)
                    try:
                        files = []
                        if row["attachments"]:
                            files = json.loads(row["attachments"])
                        self.kanban_attachments_var.set(", ".join(os.path.basename(f) for f in files) if files else "No attachments")
                    except Exception:
                        self.kanban_attachments_var.set("No attachments")

                    # enable action buttons
                    self.btn_edit.config(state="normal")
                    self.btn_delete.config(state="normal")
                    self.btn_done.config(state="normal")
                    self.btn_prev.config(state="normal" if self.kanban_selected_status != "Pending" else "disabled")
                    self.btn_next.config(state="normal" if self.kanban_selected_status != "Done" else "disabled")
                except Exception:
                    logger.exception("Error selecting kanban card")

            # --- double click opens editor ---
            def _open_editor(e=None, _id=tid):
                try:
                    if _id is not None:
                        self._open_edit_window(int(_id))
                except Exception:
                    logger.exception("Error opening card editor")
                # Returning "break" tells Tkinter to stop processing this event further
                return "break"

            # Make widget's own bindtags primary so local bindings get precedence on some platforms
            try:
                current_tags = wrapper.bindtags()
                wrapper.bindtags((str(wrapper),) + tuple(t for t in current_tags if t != str(wrapper)))
            except Exception:
                pass

            # Bind click and double-click on wrapper + child labels
            try:
                # debug wrappers still useful
                def _dbg_select(e=None, *_a, **_k):
                    logger.debug("Kanban single-click on task %s (widget=%s)", tid, getattr(e, "widget", None))
                    return _select_card(e)

                def _dbg_open(e=None, *_a, **_k):
                    logger.debug("Kanban double-click on task %s (widget=%s)", tid, getattr(e, "widget", None))
                    res = _open_editor(e)
                    # ensure we stop further propagation (prevents global bind_all from also opening editor)
                    return "break" if res is None else res

                wrapper.bind("<Button-1>", _dbg_select, add="+")
                wrapper.bind("<Double-Button-1>", _dbg_open, add="+")
                wrapper.bind("<Double-1>", _dbg_open, add="+")

                content.bind("<Button-1>", _dbg_select, add="+")
                content.bind("<Double-Button-1>", _dbg_open, add="+")
                content.bind("<Double-1>", _dbg_open, add="+")

                lbl_title.bind("<Button-1>", _dbg_select, add="+")
                lbl_title.bind("<Double-Button-1>", _dbg_open, add="+")
                lbl_title.bind("<Double-1>", _dbg_open, add="+")

                lbl_meta.bind("<Button-1>", _dbg_select, add="+")
                lbl_meta.bind("<Double-Button-1>", _dbg_open, add="+")
                lbl_meta.bind("<Double-1>", _dbg_open, add="+")

                lbl_priority.bind("<Button-1>", _dbg_select, add="+")
                lbl_priority.bind("<Double-Button-1>", _dbg_open, add="+")
                lbl_priority.bind("<Double-1>", _dbg_open, add="+")
            except Exception:
                logger.exception("Failed to bind kanban card events")

            return wrapper

        except Exception:
            logger.exception("Unhandled error creating kanban card")
            return None



    ####------------------ Kanban Board --------------------            
    def _populate_kanban(self):
        # clear existing contents
        try:
            for status, colinfo in self.kanban_columns.items():
                inner = colinfo["frame"]
                # destroy all children inside inner
                for child in list(inner.winfo_children()):
                    child.destroy()
                self.kanban_item_map[status] = []
        except Exception:
            self.kanban_item_map = {status: [] for status in STATUSES}

        
        ###
        try:
            all_rows = self.db.fetch()
        except Exception:
            all_rows = []

        ft = (self.filter_text_var.get().strip().lower() if hasattr(self, "filter_text_var") else "").strip()
        fpri = (self.filter_priority_var.get() if hasattr(self, "filter_priority_var") else "All")
        fstat = (self.filter_status_var.get() if hasattr(self, "filter_status_var") else "All")
        fdue = (self.filter_due_var.get().strip() if hasattr(self, "filter_due_var") else "").strip()
        fshow_completed = (self.filter_show_completed_var.get() if hasattr(self, "filter_show_completed_var") else "Yes")

        def _row_matches_for_kanban(r):
            if ft:
                hay = ((r["title"] or "") + " " + (r["description"] or "")).lower()
                if ft not in hay:
                    return False
            if fpri and fpri != "All":
                if (r["priority"] or "") != fpri:
                    return False
            if fstat and fstat != "All":
                if (r["status"] or "") != fstat:
                    return False
            if fdue:
                try:
                    if (r["due_date"] or "") != fdue:
                        return False
                except Exception:
                    return False
            if fshow_completed == "No" and (r["status"] or "").lower() == "done":
                return False
            return True

        rows = [r for r in all_rows if _row_matches_for_kanban(r)]

        groups = {}
        for r in rows:
            try:
                st = (r["status"] or "").strip()
                matched = None
                for s in STATUSES:
                    if s.lower() == st.lower():
                        matched = s
                        break
                if matched is None:
                    matched = st or "Pending"
                groups.setdefault(matched, []).append(r)
            except Exception:
                continue

        today = date.today()
        for status in STATUSES:
            colinfo = self.kanban_columns.get(status)
            if not colinfo:
                continue
            inner = colinfo["frame"]

            items = groups.get(status, [])
            for r in items:
                try:
                    # create a card and add it to the column frame
                    wrapper = self._create_kanban_card(inner, r)
                    self.kanban_item_map[status].append(r["id"])
                except Exception:
                    logger.exception("Error creating kanban card")
                    continue

            # ensure smallest width/wrapping is okay
            try:
                colinfo["canvas"].update_idletasks()
                # set canvas scrollregion
                colinfo["canvas"].configure(scrollregion=colinfo["canvas"].bbox("all"))
            except Exception:
                pass

    def _kanban_select(self, event):
        lb = event.widget
        idxs = lb.curselection()
        if not idxs:
            return
        idx = idxs[0]
        status = lb.status_name
        try:
            task_id = self.kanban_item_map[status][idx]
        except Exception:
            messagebox.showwarning("Error", "Could not resolve task id for selection.")
            return

        self.kanban_selected_id = task_id
        self.kanban_selected_status = status

        cur = self.db.conn.cursor()
        cur.execute("SELECT description, progress_log, outlook_id FROM tasks WHERE id=?", (task_id,))
        row = cur.fetchone()
        if not row:
            messagebox.showwarning("Error", "Task not found in database.")
            return
        desc = row["description"] or ""
        prog = row["progress_log"] or ""
        outlook_id = row["outlook_id"]

        if outlook_id and HAS_HTML:
            clean = desc or ""
            clean = re.sub(r'<style.*?>.*?</style>', '', clean, flags=re.DOTALL | re.IGNORECASE)
            clean = re.sub(r'<font[^>]*>', '', clean, flags=re.IGNORECASE).replace("</font>", "")
            clean = re.sub(r'style="[^"]*font-size:[^";]*;?"', '', clean, flags=re.IGNORECASE)
            clean = re.sub(r'style="[^"]*font-family:[^";]*;?"', '', clean, flags=re.IGNORECASE)
            clean = re.sub(r'<span[^>]*>', '<span>', clean, flags=re.IGNORECASE)
            if os.name == "nt":
                wrapper_style = "font-family:Segoe UI, Arial; font-size:9pt; line-height:1.3; color:#333;"
            else:
                wrapper_style = "font-family:Arial; font-size:11px; line-height:1.3; color:#333;"
            clean = f"<div style='{wrapper_style}'>{clean}</div>"
            try:
                self.kanban_text.pack_forget()
            except Exception:
                pass
            self.kanban_html.set_html(clean)
            self.kanban_html.pack(fill=tk.BOTH, expand=True)
        else:
            try:
                if HAS_HTML:
                    self.kanban_html.pack_forget()
            except Exception:
                pass
            self.kanban_text.delete("1.0", tk.END)
            self.kanban_text.insert(tk.END, desc)
            self.kanban_text.pack(fill=tk.BOTH, expand=True)

        self.kanban_progress.delete("1.0", tk.END)
        self.kanban_progress.insert(tk.END, prog)

        cur.execute("SELECT attachments FROM tasks WHERE id=?", (task_id,))
        row2 = cur.fetchone()
        if row2 and row2["attachments"]:
            files = json.loads(row2["attachments"])
            self.kanban_attachments_var.set(", ".join(os.path.basename(f) for f in files))
        else:
            self.kanban_attachments_var.set("No attachments")

        self.btn_edit.config(state="normal")
        self.btn_delete.config(state="normal")
        self.btn_done.config(state="normal")
        self.btn_prev.config(state="normal" if self.kanban_selected_status != "Pending" else "disabled")
        self.btn_next.config(state="normal" if self.kanban_selected_status != "Done" else "disabled")

    def _edit_selected_kanban(self):
        if not self.kanban_selected_id:
            return
        self._open_edit_window(self.kanban_selected_id)
    ###
    def _delete_selected_kanban(self):
        # Prefer selected card id
        if getattr(self, "kanban_selected_id", None):
            task_id = self.kanban_selected_id
            confirm = messagebox.askyesno("Confirm Delete", "Move selected task to Trash?")
            if not confirm:
                return
            try:
                self.db.soft_delete(task_id)
            except Exception:
                logger.exception("Soft-delete error (kanban single)")
            # clear selection UI highlight
            try:
                prev = getattr(self, "_kanban_highlighted", None)
                if prev:
                    prev.configure(highlightthickness=0)
                    self._kanban_highlighted = None
            except Exception:
                pass
            self.kanban_selected_id = None
            self.kanban_selected_status = None
            self._populate(); self._populate_kanban(); self._populate_trash()
            return

        # Fallback — if someone left old Listbox-based kanban in place
        if hasattr(self, "kanban_lists"):
            for status, lb in getattr(self, "kanban_lists", {}).items():
                sel = list(lb.curselection())
                if not sel:
                    continue
                confirm = messagebox.askyesno("Confirm Delete", f"Move {len(sel)} selected task(s) to Trash?")
                if not confirm:
                    return
                for idx in sorted(sel, reverse=True):
                    try:
                        task_id = self.kanban_item_map[status][idx]
                    except Exception:
                        continue
                    try:
                        self.db.soft_delete(task_id)
                    except Exception:
                        logger.exception("Soft-delete error (kanban)")
                    try:
                        lb.delete(idx)
                    except Exception:
                        pass
                    try:
                        del self.kanban_item_map[status][idx]
                    except Exception:
                        pass
        self._populate(); self._populate_kanban(); self._populate_trash()

    def _move_prev_selected(self):
        if not self.kanban_selected_id:
            return
        idx = STATUSES.index(self.kanban_selected_status)
        if idx > 0:
            self._move_task(self.kanban_selected_id, STATUSES[idx - 1])

    def _move_next_selected(self):
        if not self.kanban_selected_id:
            return
        idx = STATUSES.index(self.kanban_selected_status)
        if idx < len(STATUSES) - 1:
            self._move_task(self.kanban_selected_id, STATUSES[idx + 1])

    def _move_task(self, task_id, new_status):
        cur = self.db.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE id=?", (task_id,))
        r = cur.fetchone()
        if not r:
            return
        self.db.update(task_id, r["title"], r["description"], r["due_date"], r["priority"], new_status)
        self._populate()
        self._populate_kanban()
        self._sync_outlook_task(task_id, {"status": new_status}, action="update")

    def _update_progress(self):
        if not self.kanban_selected_id:
            return
        new_line = self.kanban_progress.get("1.0", tk.END).strip()
        if not new_line:
            return
        now = date.today().isoformat()
        entry = f"[{now}] {new_line}\n"
        cur = self.db.conn.cursor()
        cur.execute("SELECT progress_log FROM tasks WHERE id=?", (self.kanban_selected_id,))
        old = cur.fetchone()[0] or ""
        new_log = entry + old
        self.db.update_progress(self.kanban_selected_id, new_log)
        self.kanban_progress.delete("1.0", tk.END)
        self.kanban_progress.insert(tk.END, new_log)
        self._populate_kanban()

    # -------------------- Outlook integration --------------------
    def _send_reminder_email(self, task_id, to_address, subject_title, html_body):
        """
        Send an HTML reminder email via Outlook to `to_address`.
        Behavior:
         - If a previous reminder was sent and we stored its EntryID (reminder_mail_entryid),
           reply to that message (Reply()) so Outlook keeps the conversation/thread.
         - Otherwise create a new mail using subject_title (no extra "Reminder:" prefix)
           and set ConversationTopic when possible, then store its EntryID for future replies.
        Returns True on success, False otherwise.
        """
        if not HAS_OUTLOOK:
            logger.debug("Outlook not available; cannot send reminder email.")
            return False

        try:
            ol_app = win32com.client.Dispatch("Outlook.Application")
            ns = ol_app.GetNamespace("MAPI")
        except Exception:
            logger.exception("Failed to initialize Outlook COM objects")
            return False

        # Normalize subject: keep exactly the conversation subject you want threaded
        conv_subject = (subject_title or "Reminder").strip()

        # Attempt to read stored reminder EntryID
        entry_id = None
        try:
            if task_id:
                cur = self.db.conn.cursor()
                cur.execute("SELECT reminder_mail_entryid FROM tasks WHERE id=?", (task_id,))
                row = cur.fetchone()
                if row and row["reminder_mail_entryid"]:
                    entry_id = row["reminder_mail_entryid"]
        except Exception:
            logger.exception("Could not read reminder_mail_entryid")

        # Helper to get signature HTML using a probe item (best-effort)
        def _get_signature():
            try:
                probe = ol_app.CreateItem(0)
                probe.Display(False)
                sig = probe.HTMLBody or ""
                try:
                    probe.Close(0)
                except Exception:
                    pass
                return sig
            except Exception:
                return ""

        signature_html = _get_signature()

        # If we have an entry_id, try to get that item and reply to it (keeps thread)
        if entry_id:
            try:
                orig_item = ns.GetItemFromID(entry_id)
            except Exception:
                orig_item = None

            if orig_item is not None:
                try:
                    # Use Reply() to preserve the thread/topic; change to ReplyAll() to include all recipients.
                    reply = orig_item.Reply()
                    # Prepend our HTML content but preserve the quoted body that Reply() provides.
                    try:
                        reply.HTMLBody = (html_body or "") + signature_html + (reply.HTMLBody or "")
                    except Exception:
                        reply.Body = (html_body or "") + "\n\n" + (reply.Body or "")
                    reply.Send()

                    # After sending a reply, EntryID might change (a sent item has an EntryID).
                    try:
                        new_eid = getattr(reply, "EntryID", None)
                        if new_eid and task_id:
                            try:
                                self.db.conn.execute("UPDATE tasks SET reminder_mail_entryid=? WHERE id=?", (new_eid, task_id))
                                self.db.conn.commit()
                            except Exception:
                                logger.exception("Could not update reminder_mail_entryid after reply")
                    except Exception:
                        pass

                    logger.info("Replied to existing reminder mail for task %s -> %s", task_id, to_address)
                    return True
                except Exception:
                    logger.exception("Failed to reply to existing mail; will fallback to creating a new mail")
                    # fall through to create new mail

        # No existing thread (or reply failed) => create a new mail and ensure subject matches conv_subject
        try:
            mail = ol_app.CreateItem(0)  # olMailItem
            mail.To = to_address or ""
            # Use the exact conversation subject (no extra "Reminder:" prefix)
            mail.Subject = conv_subject
            # Try to set ConversationTopic to help Outlook thread grouping (best-effort)
            try:
                # Some Outlook object models expose ConversationTopic; set if present
                try:
                    setattr(mail, "ConversationTopic", conv_subject)
                except Exception:
                    # If attribute can't be set, ignore silently
                    pass
            except Exception:
                pass

            # Display briefly to get signature then send
            try:
                mail.Display(False)
                base_sig = mail.HTMLBody or ""
                mail.HTMLBody = (html_body or "") + base_sig
                mail.Send()
            except Exception:
                # fallback: send without signature if display failed
                try:
                    mail.HTMLBody = html_body or ""
                    mail.Send()
                except Exception:
                    logger.exception("Failed to send new reminder mail")
                    return False

            # Store EntryID on the task for future replies (if we have a real task row)
            try:
                eid = getattr(mail, "EntryID", None)
                if eid and task_id:
                    try:
                        self.db.conn.execute("UPDATE tasks SET reminder_mail_entryid=? WHERE id=?", (eid, task_id))
                        self.db.conn.commit()
                    except Exception:
                        logger.exception("Failed to store reminder_mail_entryid")
            except Exception:
                pass

            logger.info("Sent new reminder mail to %s for task %s (subject=%s)", to_address, task_id, conv_subject)
            return True

        except Exception:
            logger.exception("Failed to create/send Outlook mail")
            return False

    def _get_flagged_from_folder(self, folder, flagged):
        """Recursively fetch flagged mails from a folder + subfolders"""
        try:
            items = folder.Items
            items.Sort("[ReceivedTime]", True)
            flagged_items = items.Restrict("[FlagStatus] = 2")
            for item in flagged_items:
                if getattr(item, "Class", 0) == 43:  # MailItem
                    attachments = []
                    try:
                        if item.Attachments.Count > 0:
                            os.makedirs("attachments", exist_ok=True)
                            for att in item.Attachments:
                                fname = os.path.join("attachments", att.FileName)
                                att.SaveAsFile(fname)
                                attachments.append(fname)
                    except Exception:
                        logger.exception("Attachment import error")

                    due = None
                    try:
                        due = item.TaskDueDate.strftime("%Y-%m-%d") if getattr(item, "TaskDueDate", None) else None
                    except Exception:
                        due = None
                    desc = getattr(item, "HTMLBody", "") or getattr(item, "Body", "")
                    flagged.append({
                        "title": f"[Mail] {item.Subject}",
                        "description": desc,
                        "due_date": due,
                        "priority": "Medium",
                        "status": "Pending",
                        "outlook_id": item.EntryID,
                        "attachments": json.dumps(attachments)
                    })
            for sub in folder.Folders:
                self._get_flagged_from_folder(sub, flagged)
        except Exception:
            logger.exception("Error scanning folder for flagged items")

    def _get_flagged_emails(self):
        if not HAS_OUTLOOK:
            return []
        flagged = []
        try:
            outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            try:
                todo_folder = outlook.GetDefaultFolder(28)
                for item in todo_folder.Items:
                    if getattr(item, "Class", 0) == 43 and getattr(item, "FlagStatus", 0) == 2:
                        due = None
                        try:
                            due = item.DueDate.strftime("%Y-%m-%d") if getattr(item, "DueDate", None) else None
                        except Exception:
                            due = None
                        flagged.append({
                            "title": f"[OM] {item.Subject}",
                            "description": item.Body or "",
                            "due_date": due,
                            "priority": "Medium",
                            "status": "Pending",
                            "outlook_id": item.EntryID
                        })
            except Exception:
                logger.exception("To-Do List fetch error")

            try:
                inbox = outlook.GetDefaultFolder(6)
                self._get_flagged_from_folder(inbox, flagged)
            except Exception:
                logger.exception("Inbox flagged mail fetch error")

            try:
                search_root = outlook.GetDefaultFolder(23)
                for folder in search_root.Folders:
                    if folder.Name.lower() == "for follow up":
                        self._get_flagged_from_folder(folder, flagged)
            except Exception:
                logger.exception("Search folder fetch error")
        except Exception:
            logger.exception("Outlook fetch error")
        return flagged

    def _import_outlook_flags(self):
        flagged = self._get_flagged_emails()
        if not flagged:
            messagebox.showinfo("Outlook", "No active tasks or flagged emails found.")
            return
        cur = self.db.conn.cursor()
        new_items = [f for f in flagged if not cur.execute("SELECT 1 FROM tasks WHERE outlook_id=?", (f["outlook_id"],)).fetchone()]
        if new_items:
            self.db.bulk_add(new_items)
            self._populate(); self._populate_kanban()
        messagebox.showinfo("Outlook", f"Imported {len(new_items)} new tasks.")

    def _refresh_outlook_flags(self):
        self._import_outlook_flags()

    def _schedule_outlook_refresh(self, minutes):
        try:
            self.after(minutes * 60 * 1000, self._refresh_outlook_flags)
        except Exception:
            pass

    def _sync_outlook_task(self, task_id, data, action="update"):
        if not HAS_OUTLOOK:
            return
        try:
            outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            cur = self.db.conn.cursor()
            cur.execute("SELECT outlook_id FROM tasks WHERE id=?", (task_id,))
            row = cur.fetchone()
            if not row or not row["outlook_id"]:
                return
            entryid = row["outlook_id"]
            item = None
            try:
                todo_folder = outlook.GetDefaultFolder(28)
                for i in todo_folder.Items:
                    if i.EntryID == entryid:
                        item = i
                        break
            except Exception:
                pass
            if not item:
                try:
                    inbox = outlook.GetDefaultFolder(6)
                    item = inbox.Items.Find(f"[EntryID] = '{entryid}'")
                except Exception:
                    pass
            if not item:
                return
            if action == "done" or (action == "update" and data.get("status") == "Done"):
                if getattr(item, "Class", 0) == 48:
                    item.MarkComplete()
                elif getattr(item, "Class", 0) == 43:
                    item.FlagStatus = 1
                    item.Categories = "Completed"
                item.Save()
            elif action == "delete":
                item.Delete()
        except Exception:
            logger.exception("Outlook sync error")

    # -------------------- CSV / Contacts --------------------
    def _import_csv(self):
        path = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
        if not path:
            return
        rows = []
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    if not r.get("title"):
                        continue
                    rows.append({"title": r["title"], "description": r.get("description", ""), "due_date": r.get("due_date"),
                                 "priority": r.get("priority", "Medium"), "status": r.get("status", "Pending")})
            if rows:
                self.db.bulk_add(rows)
                self._populate(); self._populate_kanban()
                messagebox.showinfo("CSV Import", f"Imported {len(rows)} tasks.")
        except Exception:
            logger.exception("CSV import failed")
            messagebox.showerror("CSV Import", "Failed to import CSV")

    def _export_csv(self):
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if not path:
            return
        rows = self.db.fetch()
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["title", "description", "due_date", "priority", "status"])
                for r in rows:
                    writer.writerow([r["title"], r["description"], r["due_date"], r["priority"], r["status"]])
            messagebox.showinfo("CSV Export", f"Exported {len(rows)} tasks.")
        except Exception:
            logger.exception("CSV export failed")
            messagebox.showerror("CSV Export", "Export failed")

    def _import_contacts(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv")])
        if not path:
            return
        try:
            added = self.db.bulk_add_contacts_from_file(path)
            if added:
                messagebox.showinfo("Import Contacts", f"Imported {added} new contacts.")
            else:
                messagebox.showinfo("Import Contacts", "No new contacts found or already imported.")
        except Exception:
            logger.exception("Contact import error")
            messagebox.showerror("Import Contacts", "Could not import contacts")

    # -------------------- Kanban Save Desc --------------------
    def _save_kanban_desc(self):
        if not self.kanban_selected_id:
            messagebox.showwarning("No Task", "Please select a task in Kanban first.")
            return
        cur = self.db.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE id=?", (self.kanban_selected_id,))
        r = cur.fetchone()
        if not r:
            return
        if r["outlook_id"]:
            messagebox.showinfo("Info", "Outlook tasks cannot be edited here. Update directly in Outlook.")
            return
        new_desc = self.kanban_text.get("1.0", tk.END).strip()
        self.db.update(self.kanban_selected_id, r["title"], new_desc, r["due_date"], r["priority"], r["status"])
        self._populate(); self._populate_kanban()
        self._sync_outlook_task(self.kanban_selected_id, {"desc": new_desc}, action="update")
        messagebox.showinfo("Saved", "Description updated successfully.")

    # -------------------- Overdue / Today popups --------------------
    def _show_overdue_popup(self):
        rows = self.db.fetch_overdue()
        win = tk.Toplevel(self); win.title("Overdue Tasks"); win.geometry("800x400")
        if not rows:
            tk.Label(win, text="✅ No overdue tasks!", font=("", 12, "bold")).pack(padx=20, pady=20)
            return
        cols = ["id", "Title", "Due Date", "Priority", "Status"]
        tree = ttk.Treeview(win, columns=cols, show="headings", height=15)
        tree.heading("Title", text="Title")
        tree.heading("Due Date", text="Due Date")
        tree.heading("Priority", text="Priority")
        tree.heading("Status", text="Status")
        tree.heading("id", text="ID")
        tree.column("id", width=0, stretch=False, anchor="center")
        tree.column("Title", width=int(800*0.6), anchor="w")
        tree.column("Due Date", width=int(800*0.14), anchor="center")
        tree.column("Priority", width=int(800*0.13), anchor="center")
        tree.column("Status", width=int(800*0.13), anchor="center")
        for r in rows:
            tree.insert("", tk.END, values=(r["id"], r["title"], r["due_date"], r["priority"], r["status"]))
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        tree.bind("<Double-1>", lambda e, w=tree: self._open_task_on_doubleclick(e, w))

    def _show_today_popup(self):
        rows = self.db.fetch_due_today()
        win = tk.Toplevel(self); win.title("Today's Tasks"); win.geometry("800x400")
        if not rows:
            tk.Label(win, text="🎉 No tasks due today!", font=("", 12, "bold")).pack(padx=20, pady=20)
            return
        cols = ["id", "Title", "Due Date", "Priority", "Status"]
        tree = ttk.Treeview(win, columns=cols, show="headings", height=15)
        tree.heading("Title", text="Title")
        tree.heading("Due Date", text="Due Date")
        tree.heading("Priority", text="Priority")
        tree.heading("Status", text="Status")
        tree.heading("id", text="ID")
        tree.column("id", width=0, stretch=False, anchor="center")
        tree.column("Title", width=int(800*0.6), anchor="w")
        tree.column("Due Date", width=int(800*0.14), anchor="center")
        tree.column("Priority", width=int(800*0.13), anchor="center")
        tree.column("Status", width=int(800*0.13), anchor="center")
        for r in rows:
            tree.insert("", tk.END, values=(r["id"], r["title"], r["due_date"], r["priority"], r["status"]))
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        tree.bind("<Double-1>", lambda e, w=tree: self._open_task_on_doubleclick(e, w))

    # -------------------- Settings --------------------
    def _open_settings(self):
        win = tk.Toplevel(self)
        win.title("Settings")
        win.geometry("350x200")
        tk.Label(win, text="Outlook Refresh Minutes").grid(row=0, column=0, sticky="w", padx=10, pady=5)
        refresh_var = tk.IntVar(value=self.settings.get("outlook_refresh_minutes", 30))
        tk.Entry(win, textvariable=refresh_var, width=10).grid(row=0, column=1, padx=10, pady=5)
        show_desc_var = tk.BooleanVar(value=self.settings.get("show_description", False))
        tk.Checkbutton(win, text="Show Description in Task List", variable=show_desc_var).grid(row=1, column=0, columnspan=2, sticky="w", padx=10, pady=5)

        def save_and_close():
            self.settings["outlook_refresh_minutes"] = refresh_var.get()
            self.settings["show_description"] = show_desc_var.get()
            save_settings(self.settings)
            messagebox.showinfo("Settings", "Settings saved.\nRestart app to apply Task List layout changes.")
            win.destroy()

        ttk.Button(win, text="Save", command=save_and_close).grid(row=2, column=0, columnspan=2, pady=15)

    # -------------------- Trash / Delete / Restore --------------------
    def _populate_trash(self):
        try:
            for iid in getattr(self, "trash_tree", tk.Frame()).get_children():
                self.trash_tree.delete(iid)
        except Exception:
            pass
        try:
            rows = self.db.fetch_deleted()
        except Exception:
            rows = []

        # Apply same global filters to Trash view
        ft = (self.filter_text_var.get().strip().lower() if hasattr(self, "filter_text_var") else "").strip()
        fpri = (self.filter_priority_var.get() if hasattr(self, "filter_priority_var") else "All")
        fstat = (self.filter_status_var.get() if hasattr(self, "filter_status_var") else "All")
        fdue = (self.filter_due_var.get().strip() if hasattr(self, "filter_due_var") else "").strip()

        def _trash_row_matches(r):
            if ft:
                hay = ((r["title"] or "") + " " + (r["title"] or "")).lower()
                if ft not in hay:
                    return False
            if fpri and fpri != "All":
                if (r["priority"] or "") != fpri:
                    return False
            if fstat and fstat != "All":
                if (r["status"] or "") != fstat:
                    return False
            if fdue:
                try:
                    if (r["due_date"] or "") != fdue:
                        return False
                except Exception:
                    return False
            return True

        rows = [r for r in rows if _trash_row_matches(r)]

        for r in rows:
            try:
                deleted_at = r["deleted_at"] or "?"
                self.trash_tree.insert("", tk.END, values=(r["id"], r["title"], deleted_at, r["due_date"] or "—", r["priority"], r["status"]))
            except Exception:
                pass

    def _restore_selected_trash(self):
        sel = self.trash_tree.selection()
        if not sel:
            return
        for s in sel:
            try:
                task_id = int(self.trash_tree.item(s, "values")[0])
                self.db.restore(task_id)
            except Exception:
                logger.exception("Restore error")
        self._populate()
        self._populate_kanban()
        self._populate_trash()

    def _permanently_delete_selected_trash(self):
        sel = self.trash_tree.selection()
        if not sel:
            return
        confirm = messagebox.askyesno("Confirm Permanent Delete", f"Permanently delete {len(sel)} selected item(s)? This cannot be undone.")
        if not confirm:
            return
        for s in sel:
            try:
                task_id = int(self.trash_tree.item(s, "values")[0])
                try:
                    cur = self.db.conn.cursor()
                    cur.execute("SELECT outlook_id FROM tasks WHERE id=?", (task_id,))
                    row = cur.fetchone()
                    if row and row["outlook_id"]:
                        try:
                            self._sync_outlook_task(task_id, {}, action="delete")
                        except Exception:
                            pass
                except Exception:
                    pass
                self.db.delete(task_id)
            except Exception:
                logger.exception("Permanent delete error")
        self._populate()
        self._populate_kanban()
        self._populate_trash()

    def _empty_trash_confirm(self):
        confirm = messagebox.askyesno("Empty Trash", "Permanently delete all items in Trash? This cannot be undone.")
        if not confirm:
            return
        try:
            rows = self.db.fetch_deleted()
            for r in rows:
                try:
                    if r["outlook_id"]:
                        try:
                            self._sync_outlook_task(r["id"], {}, action="delete")
                        except Exception:
                            pass
                except Exception:
                    pass
            self.db.purge_deleted()
        except Exception:
            logger.exception("Empty trash error")
        self._populate()
        self._populate_kanban()
        self._populate_trash()

    def _delete_task(self):
        sel = self.tree.selection()
        if not sel:
            return
        confirm = messagebox.askyesno("Confirm Delete", f"Move {len(sel)} selected task(s) to Trash?")
        if not confirm:
            return
        for s in sel:
            try:
                task_id = int(self.tree.item(s, "values")[0])
                self.db.soft_delete(task_id)
            except Exception:
                logger.exception("Soft-delete error")
        self._populate()
        self._populate_kanban()
        try:
            self._populate_trash()
        except Exception:
            pass

    def _mark_done(self):
        sel = self.tree.selection()
        if not sel:
            return
        for s in sel:
            try:
                task_id = int(self.tree.item(s, "values")[0])
            except Exception:
                continue
            cur = self.db.conn.cursor()
            cur.execute("SELECT * FROM tasks WHERE id=?", (task_id,))
            row = cur.fetchone()
            if not row:
                continue
            self.db.mark_done(task_id)
            try:
                self._create_next_occurrence_if_needed(row)
            except Exception:
                pass
            self._sync_outlook_task(task_id, {}, action="done")
        self._populate(); self._populate_kanban()

    def _mark_done_selected_kanban(self):
        if not self.kanban_selected_id:
            return
        cur = self.db.conn.cursor()
        cur.execute("SELECT * FROM tasks WHERE id=?", (self.kanban_selected_id,))
        row = cur.fetchone()
        if not row:
            return
        self.db.mark_done(self.kanban_selected_id)
        try:
            self._create_next_occurrence_if_needed(row)
        except Exception:
            pass
        self._populate(); self._populate_kanban()
        self._sync_outlook_task(self.kanban_selected_id, {}, action="done")

    def _create_next_occurrence_if_needed(self, task_row):
        try:
            rec = (task_row["recurrence"] or "none")
            due = task_row["due_date"]
            if not due:
                return None
            if not rec or rec.strip().lower() in ("none", ""):
                return None
            next_due = self._compute_next_due(due, rec)
            if not next_due:
                return None
            attachments = task_row["attachments"] or None
            progress = task_row["progress_log"] or None
            self.db.add(
                title=task_row["title"],
                description=task_row["description"],
                due_date=next_due,
                priority=task_row["priority"] or "Medium",
                status="Pending",
                outlook_id=None,
                reminder_minutes=task_row["reminder_minutes"],
                reminder_set_at=None,
                recurrence=rec
            )
            cur = self.db.conn.cursor()
            cur.execute("SELECT last_insert_rowid() as id")
            new_id = cur.fetchone()["id"]
            if attachments:
                self.db.conn.execute("UPDATE tasks SET attachments=? WHERE id=?", (attachments, new_id))
            if progress:
                self.db.update_progress(new_id, progress)
            return new_id
        except Exception:
            logger.exception("Error creating next recurrence")
            return None

    # -------------------- Misc --------------------
    def _on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            self.attachments_var.set("")
            return
        vals = self.tree.item(sel[0], "values")
        task_id = int(vals[0])
        cur = self.db.conn.cursor()
        cur.execute("SELECT attachments FROM tasks WHERE id=?", (task_id,))
        row = cur.fetchone()
        if row and row["attachments"]:
            try:
                files = json.loads(row["attachments"])
            except Exception:
                files = []
        else:
            files = []
        if hasattr(self, "attachments_var"):
            try:
                self.attachments_var.set(", ".join(os.path.basename(f) for f in files))
            except Exception:
                self.attachments_var.set("")

    def _on_delete_key(self, event=None):
        # Only trigger when Delete key is pressed (prevent BackSpace from acting)
        try:
            if event is not None:
                keysym = getattr(event, "keysym", None)
                # On some platforms event may be None (when called directly) so allow None -> treat as Delete
                if keysym is not None and keysym != "Delete":
                    return
        except Exception:
            pass

        try:
            idx = self.notebook.index(self.notebook.select())
            tab_text = self.notebook.tab(idx, "text")
        except Exception:
            tab_text = None
        if tab_text == "Task List":
            self._delete_task()
        elif tab_text == "Kanban Board":
            self._delete_selected_kanban()
        elif tab_text == "Trash":
            self._permanently_delete_selected_trash()
        else:
            self._delete_task()

    def _treeview_sort_column(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        try:
            l.sort(key=lambda t: datetime.strptime(t[0], "%Y-%m-%d"), reverse=reverse)
        except Exception:
            l.sort(reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, "", index)
        self.tree.heading(col, command=lambda: self._treeview_sort_column(col, not reverse))

    def _check_reminders(self):
        due_today = self.db.fetch_due_today()
        if due_today and HAS_NOTIFY:
            _safe_show_toast("Tasks Due Today", f"{len(due_today)} tasks due today")
        self.after(3600 * 1000, self._check_reminders)

    def _on_exit(self):
        try:
            if hasattr(self, "db") and getattr(self.db, "conn", None):
                try:
                    self.db.conn.close()
                except Exception:
                    pass
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            try:
                sys.exit(0)
            except Exception:
                pass


# -------------------- Main --------------------
def main():
    app = TaskApp()
    app.mainloop()


if __name__ == "__main__":
    main()